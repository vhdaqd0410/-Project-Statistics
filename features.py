# -*- coding: utf-8 -*-
"""AI后期剪辑提成工具 - 扩展功能模块 v2.0"""
import os, re, json, shutil, datetime, html as html_mod, random
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ===================== 工具函数 =====================

def _find_chinese_font():
    """自动查找系统中可用的中文字体"""
    candidates = [
        r'C:\Windows\Fonts\simsun.ttc',
        r'C:\Windows\Fonts\msyh.ttc',
        r'C:\Windows\Fonts\msyhbd.ttc',
        r'C:\Windows\Fonts\simkai.ttf',
    ]
    for f in candidates:
        if os.path.exists(f):
            return f
    return None


def _escape_html(text):
    """HTML转义"""
    return html_mod.escape(str(text))


# ===================== 1. 生成下月模板 =====================

def generate_next_month_template(template_path, output_dir=None):
    if output_dir is None:
        output_dir = os.path.dirname(template_path)
    CN_MONTHS = ['', '一月','二月','三月','四月','五月','六月',
                 '七月','八月','九月','十月','十一月','十二月']
    wb = load_workbook(template_path)
    ws = wb.active

    title_cell = None
    for col in [2, 1]:
        val = ws.cell(1, col).value
        if val and re.search(r'\d+年\d+月', str(val)):
            title_cell = (1, col)
            break
    if not title_cell:
        return None, "无法识别模板中的年月标题"

    old_title = str(ws.cell(*title_cell).value)
    m = re.search(r'(\d+)年(\d+)月', old_title)
    if not m:
        return None, "无法解析月份"

    year, month = int(m.group(1)), int(m.group(2))
    month += 1
    if month > 12:
        month = 1; year += 1

    new_title = re.sub(r'\d+年\d+月', f'{year}年{month:02d}月', old_title)
    ws.cell(*title_cell, new_title)

    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= 4:
            ws.unmerge_cells(str(mc))
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
        for cell in row:
            try: cell.value = None
            except AttributeError: pass

    cn = CN_MONTHS[month]
    output = os.path.join(output_dir, f'AI后期剪辑提成一组模板_{year}年{month:02d}月.xlsx')
    wb.save(output); wb.close()
    return output, f"下月模板已生成：{year}年{month:02d}月（{cn}）"


# ===================== 2. 导出PDF =====================

def export_to_pdf(excel_path, pdf_path=None):
    if pdf_path is None:
        pdf_path = excel_path.replace('.xlsx', '.pdf')

    font_path = _find_chinese_font()
    if not font_path:
        raise RuntimeError("未找到可用中文字体，PDF导出失败")

    from fpdf import FPDF
    wb = load_workbook(excel_path)
    ws = wb.active

    pdf = FPDF(orientation='L', unit='mm', format='A3')
    pdf.add_font('CJK', '', font_path, uni=True)
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    col_widths = [8, 12, 12, 5, 8, 10, 18, 10, 10, 15, 8, 8, 8, 8, 8, 8, 15, 8, 6, 12]
    headers = [str(ws.cell(3, c).value)[:6] if ws.cell(3, c).value else '' for c in range(1, 21)]

    def draw_header(y):
        for i, hdr in enumerate(headers):
            x = 5 + sum(col_widths[:i])
            pdf.set_xy(x, y)
            pdf.set_font('CJK', '', 5)
            pdf.cell(col_widths[i], 6, hdr, border=1, align='C')

    draw_header(10)
    row_y = 16
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
        if row_y > 280:
            pdf.add_page(); row_y = 10
            draw_header(row_y); row_y += 6
        for ci, cell in enumerate(row[:20]):
            v = str(cell.value)[:12] if cell.value is not None else ''
            x = 5 + sum(col_widths[:ci])
            pdf.set_xy(x, row_y)
            pdf.set_font('CJK', '', 5)
            pdf.cell(col_widths[ci], 6, v, border=1, align='C')
        row_y += 6

    pdf.output(pdf_path); wb.close()
    return pdf_path


# ===================== 3. 个人绩效卡片 =====================

def generate_person_cards(records, commission_data, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    persons = {}
    for r in records:
        nm = r['身份证姓名']
        if nm not in persons:
            persons[nm] = {'projects': [], 'total_eps': 0}
        persons[nm]['projects'].append(r)
        persons[nm]['total_eps'] += r['单项目数/集数']

    cards = []
    for nm in sorted(persons.keys()):
        cd = commission_data.get(nm, {})
        proj_rows = ''
        for p in persons[nm]['projects']:
            pid = _escape_html(str(p['项目ID']))
            pname = _escape_html(str(p['AI项目名称'])[:30])
            d = p['结束日期'].strftime('%m/%d') if p['结束日期'] else ''
            proj_rows += f'<tr><td>{pid}</td><td>{pname}</td><td>{p["单项目数/集数"]}集</td><td>{d}</td></tr>'

        sc = '#27ae60' if cd.get('is_complete') == '是' else '#e74c3c'
        st = '达标' if cd.get('is_complete') == '是' else '未达标'
        q = cd.get('quota', 0)
        qt = f'基准{q}集' if q > 0 else '无基准'
        tc = cd.get('total_commission', 0)
        cc = '#27ae60' if tc >= 0 else '#e74c3c'

        card = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>{_escape_html(nm)} - 绩效卡片</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px}}
.card{{max-width:600px;margin:0 auto;background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,.1);overflow:hidden}}
.hd{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);color:#fff;padding:24px 28px;text-align:center}}
.hd h1{{font-size:22px}}.hd .role{{font-size:13px;opacity:.85}}
.stats{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;padding:20px 24px}}
.stat{{text-align:center;padding:12px 8px;background:#f8fafc;border-radius:10px}}
.stat .v{{font-size:28px;font-weight:700}}.stat .l{{font-size:11px;color:#64748b;margin-top:2px}}
.status{{display:inline-block;padding:3px 12px;border-radius:12px;font-size:13px;font-weight:600;color:#fff;background:{sc}}}
.ps{{padding:0 24px 20px}}.ps h3{{font-size:14px;color:#334155;margin-bottom:8px}}
.ps table{{width:100%;border-collapse:collapse;font-size:12px}}
.ps th{{background:#e2e8f0;padding:6px 8px;text-align:left}}
.ps td{{padding:5px 8px;border-bottom:1px solid #e2e8f0}}
.ft{{background:#f8fafc;padding:16px 24px;text-align:center;font-size:12px;color:#94a3b8}}
</style></head><body><div class="card">
<div class="hd"><h1>{_escape_html(nm)}</h1><div class="role">{cd.get("role","")} · {qt}</div></div>
<div class="stats">
<div class="stat"><div class="v">{persons[nm]["total_eps"]}</div><div class="l">完成集数</div></div>
<div class="stat"><div class="v"><span class="status">{st}</span></div><div class="l">绩效状态</div></div>
<div class="stat"><div class="v" style="color:{cc}">{tc:,}</div><div class="l">提成(元)</div></div>
</div>
<div class="ps"><h3>本月项目明细 ({len(persons[nm]["projects"])}个)</h3>
<table><tr><th>项目ID</th><th>名称</th><th>集数</th><th>交付</th></tr>{proj_rows}</table></div>
<div class="ft">AI后期剪辑提成 · {datetime.date.today().strftime("%Y年%m月")}</div>
</div></body></html>'''
        cards.append((nm, card))

    paths = []
    for nm, html in cards:
        fpath = os.path.join(output_dir, f'{nm}_绩效卡片.html')
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(html)
        paths.append(fpath)

    idx_html = '''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>全员绩效卡片</title>
<style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px}h1{text-align:center;color:#1e3a5f;margin-bottom:20px}.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:16px;max-width:1400px;margin:0 auto}</style>
</head><body><h1>AI后期剪辑 · 全员绩效卡片</h1><div class="grid">'''
    for nm, _ in cards:
        idx_html += f'<iframe src="{_escape_html(nm)}_绩效卡片.html" style="width:100%;height:320px;border:none;border-radius:12px;background:#fff;box-shadow:0 2px 12px rgba(0,0,0,.06)"></iframe>\n'
    idx_html += '</div></body></html>'
    idx_path = os.path.join(output_dir, 'index.html')
    with open(idx_path, 'w', encoding='utf-8') as f:
        f.write(idx_html)
    paths.insert(0, idx_path)
    return paths


# ===================== 4. 多月份对比 =====================

def compare_months(file1_path, file2_path):
    wb1 = load_workbook(file1_path, data_only=True); wb2 = load_workbook(file2_path, data_only=True)
    ws1, ws2 = wb1.active, wb2.active

    m1 = re.search(r'(\d+)年(\d+)月', str(ws1.cell(1, 2).value or ws1.cell(1, 1).value or ''))
    m2 = re.search(r'(\d+)年(\d+)月', str(ws2.cell(1, 2).value or ws2.cell(1, 1).value or ''))
    label1 = f"{m1.group(1)}年{m1.group(2)}月" if m1 else "文件1"
    label2 = f"{m2.group(1)}年{m2.group(2)}月" if m2 else "文件2"

    def parse_sheet(ws):
        data = {}
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
            name = str(row[1].value).strip() if row[1].value else ''
            if not name: continue
            eps = int(row[10].value) if row[10].value else 0
            comm = int(row[17].value) if row[17].value else 0
            if name not in data: data[name] = {'eps': 0, 'comm': 0}
            data[name]['eps'] += eps
            data[name]['comm'] += comm
        return data

    d1, d2 = parse_sheet(ws1), parse_sheet(ws2)
    all_names = sorted(set(list(d1.keys()) + list(d2.keys())))
    diff_rows = []
    for nm in all_names:
        e1, e2 = d1.get(nm, {}).get('eps', 0), d2.get(nm, {}).get('eps', 0)
        c1, c2 = d1.get(nm, {}).get('comm', 0), d2.get(nm, {}).get('comm', 0)
        de, dc = e2 - e1, c2 - c1
        if de != 0 or dc != 0:
            diff_rows.append((nm, e1, e2, de, c1, c2, dc))
    wb1.close(); wb2.close()
    return label1, label2, diff_rows


# ===================== 5. 备份 =====================

def backup_output(excel_path, html_path, backup_dir):
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    backed = []
    for fp in [excel_path, html_path]:
        if fp and os.path.exists(fp):
            name, ext = os.path.splitext(os.path.basename(fp))
            dest = os.path.join(backup_dir, f'{name}_{ts}{ext}')
            shutil.copy2(fp, dest)
            backed.append(dest)
    return backed


# ===================== 6. 数据预览 =====================

def data_preview(records, commission_data):
    """返回数据预览信息：每人集数/项目数/绩效/提成"""
    persons = {}
    for r in records:
        nm = r['身份证姓名']
        if nm not in persons:
            persons[nm] = {'eps': 0, 'projects': set()}
        persons[nm]['eps'] += r['单项目数/集数']
        if r['项目ID']:
            persons[nm]['projects'].add(r['项目ID'])

    preview = []
    for nm in sorted(persons.keys()):
        cd = commission_data.get(nm, {})
        preview.append({
            'name': nm,
            'role': cd.get('role', ''),
            'episodes': persons[nm]['eps'],
            'projects': len(persons[nm]['projects']),
            'quota': cd.get('quota', 0),
            'status': cd.get('is_complete', ''),
            'commission': cd.get('total_commission', 0),
        })
    return preview


# ===================== 7. 组内排名 =====================

def generate_ranking_html(commission_data, output_path=None):
    """生成组内排名HTML"""
    persons = []
    for nm, cd in commission_data.items():
        if not cd: continue
        persons.append({
            'name': nm, 'role': cd.get('role', ''),
            'eps': cd.get('total_episodes', 0),
            'comm': cd.get('total_commission', 0),
            'status': cd.get('is_complete', ''),
        })

    by_eps = sorted(persons, key=lambda x: x['eps'], reverse=True)
    by_comm = sorted(persons, key=lambda x: x['comm'], reverse=True)

    def rank_rows(data, key):
        rows = ''
        for i, p in enumerate(data):
            medal = ['🥇','🥈','🥉'][i] if i < 3 else str(i+1)
            v = p[key]
            color = '#27ae60' if v >= 0 else '#e74c3c'
            sign = '集' if key == 'eps' else '元'
            rows += f'<tr><td>{medal}</td><td><b>{_escape_html(p["name"])}</b></td><td>{p["role"]}</td><td style="color:{color};font-weight:600">{v:,}{sign}</td><td>{"✅" if p["status"]=="是" else "❌"}</td></tr>'
        return rows

    html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>剪辑一组 · 组内排名</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px;color:#2c3e50}}
h1{{text-align:center;color:#1e3a5f;margin:20px 0}}
.ct{{max-width:1100px;margin:0 auto}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}
.card{{background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.06);overflow:hidden}}
.card h2{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);color:#fff;padding:14px 20px;font-size:16px}}
table{{width:100%;border-collapse:collapse}}
th{{background:#e2e8f0;padding:8px 12px;font-size:12px;text-align:left}}
td{{padding:8px 12px;font-size:13px;border-bottom:1px solid #e2e8f0}}
tr:hover{{background:#f8fafc}}
</style></head><body>
<h1>剪辑一组 · 月度排名</h1>
<div class="ct"><div class="row">
<div class="card"><h2>📊 集数排行</h2><table><tr><th>排名</th><th>姓名</th><th>角色</th><th>集数</th><th>绩效</th></tr>{rank_rows(by_eps, 'eps')}</table></div>
<div class="card"><h2>💰 提成排行</h2><table><tr><th>排名</th><th>姓名</th><th>角色</th><th>提成</th><th>绩效</th></tr>{rank_rows(by_comm, 'comm')}</table></div>
</div></div></body></html>'''

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
    return html


# ===================== 8. 项目管理 =====================

def generate_project_management_html(records, output_path=None):
    """生成项目管理视图HTML"""
    projects = {}
    for r in records:
        pid = r['项目ID']
        if not pid: continue
        if pid not in projects:
            projects[pid] = {
                'name': r['AI项目名称'][:60],
                'type': r['项目类型'],
                'date': r['结束日期'],
                'people': {},
                'total_eps': 0,
            }
        nm = r['身份证姓名']
        if nm not in projects[pid]['people']:
            projects[pid]['people'][nm] = 0
        projects[pid]['people'][nm] += r['单项目数/集数']
        projects[pid]['total_eps'] += r['单项目数/集数']

    sorted_projs = sorted(projects.items(), key=lambda x: x[1]['date'] or datetime.date(2000,1,1))

    rows = ''
    for i, (pid, info) in enumerate(sorted_projs):
        d = info['date'].strftime('%m/%d') if info['date'] else '-'
        people_str = ', '.join(f'{n}({e}集)' for n, e in info['people'].items())
        rows += f'''<tr>
<td>{i+1}</td><td><b>{pid}</b></td><td>{_escape_html(info["name"])}</td>
<td>{info["total_eps"]}</td><td>{len(info["people"])}人</td><td>{d}</td>
<td style="font-size:11px">{_escape_html(people_str)}</td></tr>\n'''

    html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>项目管理 · 项目清单</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px;color:#2c3e50}}
h1{{text-align:center;color:#1e3a5f;margin:20px 0}}
.ct{{max-width:1200px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.06);overflow:hidden}}
table{{width:100%;border-collapse:collapse}}
th{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);color:#fff;padding:10px 12px;font-size:12px;text-align:left}}
td{{padding:8px 12px;font-size:13px;border-bottom:1px solid #e2e8f0}}
tr:hover{{background:#f8fafc}}
.summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;padding:20px;max-width:1200px;margin:0 auto 20px}}
.s{{background:#fff;border-radius:10px;padding:16px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.05)}}
.s .v{{font-size:28px;font-weight:700;color:#1e3a5f}}
.s .l{{font-size:12px;color:#64748b}}
</style></head><body>
<h1>项目管理 · 项目清单</h1>
<div class="summary">
<div class="s"><div class="v">{len(projects)}</div><div class="l">项目总数</div></div>
<div class="s"><div class="v">{sum(p["total_eps"] for p in projects.values())}</div><div class="l">总集数</div></div>
<div class="s"><div class="v">{len(set(n for p in projects.values() for n in p["people"]))}</div><div class="l">参与人数</div></div>
<div class="s"><div class="v">{sum(len(p["people"]) for p in projects.values())}</div><div class="l">人次</div></div>
</div>
<div class="ct"><table>
<tr><th>#</th><th>项目ID</th><th>项目名称</th><th>集数</th><th>人数</th><th>交付</th><th>参与人员(集数)</th></tr>
{rows}</table></div></body></html>'''

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
    return html


# ===================== 9. 智能分集 =====================

def smart_episode_assignment(total_eps, selected_people, role_map, card1_range=15):
    """
    三阶段分集（一卡少拿、二卡/助理多拿）:
    1) 1-3集   → 小组长独享
    2) 4~card1_range → 仅一卡平分
    3) 后段       → 加权: 小组长0.3, 组长0.4, 一卡0.3, 二卡/助理1.0
    """
    leader = [p for p in selected_people if '小组长' in role_map.get(p, '')]
    card1_only = [p for p in selected_people if '一卡' in role_map.get(p, '') and p not in leader]
    card2_people = [p for p in selected_people if '二卡' in role_map.get(p, '') or '助理' in role_map.get(p, '')]
    big_leader = [p for p in selected_people if '组长' in role_map.get(p, '') and '小组长' not in role_map.get(p, '')]

    first3 = list(range(1, min(4, total_eps + 1)))
    mid_range = list(range(4, min(card1_range, total_eps) + 1)) if total_eps >= 4 else []
    tail_range = list(range(card1_range + 1, total_eps + 1)) if total_eps > card1_range else []

    result = {p: [] for p in selected_people}

    # Phase 1: 1-3 小组长
    for ln in leader:
        result[ln].extend(first3)

    # Phase 2: 4~card1_range 仅一卡平分
    if card1_only and mid_range:
        random.shuffle(card1_only)
        n = len(card1_only)
        base, rem = divmod(len(mid_range), n)
        pos = 0
        for i, p in enumerate(card1_only):
            cnt = base + (1 if i < rem else 0)
            if cnt > 0:
                result[p].extend(mid_range[pos:pos + cnt])
                pos += cnt

    # Phase 3: 后段加权 —— 一卡拿极少(0.3)，二卡/助理拿主力(1.0)
    if not tail_range:
        return _format_result(result, leader, card1_only, card2_people, big_leader,
                              selected_people, len(first3), len(mid_range), len(tail_range),
                              card1_range)

    all_p = list(selected_people)
    wmap = {}
    for p in all_p:
        if p in leader:             wmap[p] = 0.3
        elif p in big_leader:       wmap[p] = 0.4
        elif p in card1_only:       wmap[p] = 0.3   # 一卡已有4-15段，后段尽量少
        else:                       wmap[p] = 1.0   # 二卡/助理主力
    total_w = sum(wmap.values())
    total_t = len(tail_range)
    if total_w == 0 or total_t == 0:
        return _format_result(result, leader, card1_only, card2_people, big_leader,
                              selected_people, len(first3), len(mid_range), len(tail_range),
                              card1_range)

    raw = [total_t * wmap[p] / total_w for p in all_p]
    counts = [max(0, round(r)) for r in raw]
    diff = total_t - sum(counts)
    if diff != 0:
        idx_sorted = sorted(range(len(all_p)), key=lambda i: (raw[i] - counts[i]) * (-1 if diff > 0 else 1))
        for k in range(abs(diff)):
            i = idx_sorted[k]
            if diff > 0 or counts[i] > 1:
                counts[i] += (1 if diff > 0 else -1)

    indices = list(range(len(all_p)))
    random.shuffle(indices)
    pos = 0
    for idx in indices:
        cnt = counts[idx]
        if cnt > 0 and pos < total_t:
            actual = min(cnt, total_t - pos)
            result[all_p[idx]].extend(tail_range[pos:pos + actual])
            pos += actual

    return _format_result(result, leader, card1_only, card2_people, big_leader,
                          selected_people, len(first3), len(mid_range), len(tail_range),
                          card1_range)


def _format_result(result, leader, card1_only, card2_people, big_leader,
                   selected_people, n_first, n_mid, n_tail, card1_range=15):
    def merge_ranges(eps):
        if not eps: return []
        eps = sorted(set(eps))
        ranges, s, e = [], eps[0], eps[0]
        for ep in eps[1:]:
            if ep == e + 1: e = ep
            else: ranges.append((s, e)); s = e = ep
        ranges.append((s, e))
        return ranges

    assignments, fmt, summary = {}, {}, {}
    for name in result:
        r = merge_ranges(result[name])
        assignments[name] = r
        summary[name] = len(result[name])
        fmt[name] = '；'.join(f'{s}-{e}' if s != e else str(s) for s, e in r)

    stats = {
        '小组长人数': len(leader),
        '一卡人数': len(card1_only),
        '二卡助理人数': len(card2_people),
        '剪辑组长人数': len(big_leader),
        '总人数': len(selected_people),
        '前三集': n_first, f'中段(4-{card1_range})': n_mid,
        f'后段({card1_range + 1}+)': n_tail,
    }
    return {'assignments': assignments, 'summary': summary, 'formatted': fmt, 'stats': stats}


# ===================== 10. 个人月度趋势 =====================

def generate_person_trend_html(person_name, role_map, output_dir):
    """扫描output_dir下所有AI后期剪辑提成一组*xlsx，提取某人的月度集数和提成趋势"""
    pattern = re.compile(r'AI后期剪辑提成一组(.*?)\.xlsx')
    months = []
    for fname in sorted(os.listdir(output_dir)):
        m = pattern.match(fname)
        if not m or '_仪表盘' in fname: continue
        month_label = m.group(1)
        fpath = os.path.join(output_dir, fname)
        try:
            wb = load_workbook(fpath, data_only=True)
            ws = wb.active
            found_eps, found_comm = 0, 0
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
                name_cell = str(row[1].value).strip() if row[1].value else ''
                if person_name not in name_cell: continue
                if row[10].value:
                    found_eps += int(row[10].value) if str(row[10].value).strip().lstrip('-').isdigit() else 0
                if row[17].value:
                    found_comm += int(row[17].value) if str(row[17].value).strip().lstrip('-').isdigit() else 0
            wb.close()
            months.append({'month': month_label, 'eps': found_eps, 'comm': found_comm})
        except Exception:
            continue

    if not months:
        return None

    # CSS 柱状图
    max_eps = max(m['eps'] for m in months) or 1
    max_comm = max(abs(m['comm']) for m in months) or 1
    eps_bars = ''
    comm_bars = ''
    quota = 70 if '一卡' in role_map.get(person_name, '') else (120 if '二卡' in role_map.get(person_name, '') or '助理' in role_map.get(person_name, '') else 0)

    for m in months:
        ep_pct = min(100, int(m['eps'] / max_eps * 100)) if max_eps else 0
        ec = '#27ae60' if m['eps'] >= quota else '#e74c3c'
        cp = abs(m['comm']) / max_comm * 100 if max_comm else 0
        cc = '#27ae60' if m['comm'] >= 0 else '#e74c3c'
        eps_bars += f'<div class="br"><span class="bl">{m["month"]}</span><div class="bt"><div class="bf" style="width:{ep_pct}%;background:{ec}"></div></div><span class="bv">{m["eps"]}集</span></div>'
        comm_bars += f'<div class="br"><span class="bl">{m["month"]}</span><div class="bt"><div class="bf" style="width:{min(cp,100):.0f}%;background:{cc}"></div></div><span class="bv" style="color:{cc}">{m["comm"]:,}元</span></div>'

    quota_info = f'基准{quota}集' if quota > 0 else '无基准'
    html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>{_escape_html(person_name)} · 月度趋势</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px;color:#2c3e50}}
h1{{text-align:center;color:#1e3a5f;margin:20px 0}}
.ct{{max-width:900px;margin:0 auto}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:24px}}
.card{{background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.06);padding:20px}}
.card h2{{font-size:16px;margin-bottom:16px;padding-bottom:8px;border-bottom:2px solid #e2e8f0}}
.br{{display:flex;align-items:center;margin:6px 0;gap:8px}}
.bl{{width:50px;font-size:12px;text-align:right;flex-shrink:0}}
.bt{{flex:1;height:18px;background:#e2e8f0;border-radius:9px;overflow:hidden;position:relative}}
.bf{{height:100%;border-radius:9px;transition:width .3s}}
.bv{{width:70px;font-size:12px;font-weight:600;flex-shrink:0;text-align:right}}
.info{{text-align:center;color:#64748b;font-size:13px;margin:8px 0 20px}}
</style></head><body>
<h1>{_escape_html(person_name)} · 月度趋势</h1>
<div class="info">{role_map.get(person_name, '')} · {quota_info} · 共{len(months)}个月数据</div>
<div class="ct"><div class="row">
<div class="card"><h2>📊 集数变化</h2>{eps_bars}</div>
<div class="card"><h2>💰 提成变化</h2>{comm_bars}</div>
</div></div></body></html>'''
    trend_path = os.path.join(output_dir, f'{person_name}_月度趋势.html')
    with open(trend_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return trend_path


# ===================== 11. 数据校验 =====================

def validate_project_data(project_file, role_map):
    """扫描项目数据文件，返回校验问题列表"""
    import pandas as pd
    issues = []
    try:
        df = pd.read_excel(project_file, header=None)
    except Exception as e:
        return [('文件', '无法读取', str(e))]

    known_names = set(role_map.keys())
    proj_name = ''; proj_id = ''
    seen_eps = {}  # 项目ID -> 集数集合（检测重复分配）

    for idx in range(len(df)):
        col1 = str(df.iloc[idx, 0]) if pd.notna(df.iloc[idx, 0]) else ''
        col2 = str(df.iloc[idx, 1]) if pd.notna(df.iloc[idx, 1]) else ''
        col3 = str(df.iloc[idx, 2]) if pd.notna(df.iloc[idx, 2]) else ''
        col4 = str(df.iloc[idx, 3]) if pd.notna(df.iloc[idx, 3]) else ''

        if col1 not in ['', 'nan'] and '月份' not in col1:
            proj_name = col1[:60]
            m = re.search(r'\d{4}', col1)
            if m: 
                proj_id = m.group()
                if proj_id not in seen_eps:
                    seen_eps[proj_id] = set()

        if col3 not in ['', 'nan']:
            for match in re.finditer(r'([^\s：:]+)[：:]\s*(.+)', col3):
                name = match.group(1).strip()
                eps_part = match.group(2).strip()
                if name and name not in known_names:
                    issues.append((f'行{idx+1}', name, f'"{name}" 不在人员配置中'))
                # 检测集数范围
                nums = re.findall(r'\d+', eps_part)
                if nums:
                    for n in nums:
                        ep = int(n)
                        if proj_id:
                            if ep in seen_eps.get(proj_id, set()):
                                issues.append((f'行{idx+1}', name, f'集数{ep}在项目{proj_id}中已被分配过'))
                            seen_eps[proj_id].add(ep)

        # 检测日期格式
        if col4 not in ['', 'nan', '多版本已交付', '已分集', '已分发']:
            if not re.search(r'\d+\.\d+', col4):
                issues.append((f'行{idx+1}', '日期', f'日期格式异常: {col4[:20]}'))

    return issues if issues else []


# ===================== 12. 备份管理 =====================

def list_backups(backup_dir):
    """列出备份目录中所有备份"""
    if not os.path.exists(backup_dir):
        return []
    files = [f for f in os.listdir(backup_dir) if f.endswith('.xlsx') or f.endswith('.html')]
    files.sort(reverse=True)
    return [{'name': f, 'path': os.path.join(backup_dir, f),
             'size': os.path.getsize(os.path.join(backup_dir, f)),
             'mtime': datetime.datetime.fromtimestamp(os.path.getmtime(os.path.join(backup_dir, f)))}
            for f in files]

def cleanup_backups(backup_dir, keep=30):
    """清理旧备份，只保留最近keep个"""
    if not os.path.exists(backup_dir):
        return 0
    files = [f for f in os.listdir(backup_dir) if f.endswith('.xlsx') or f.endswith('.html')]
    files.sort()
    to_remove = max(0, len(files) - keep)
    for f in files[:to_remove]:
        os.remove(os.path.join(backup_dir, f))
    return to_remove
