# -*- coding: utf-8 -*-
"""AI后期剪辑提成表生成工具 - GUI v6.0"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess, threading, os, sys, json, re, tempfile
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CLI_SCRIPT = os.path.join(SCRIPT_DIR, 'generate_commission.py')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'config.json')
BACKUP_DIR = os.path.join(SCRIPT_DIR, 'backup')
CARDS_DIR = os.path.join(SCRIPT_DIR, '个人绩效卡片')

# 导入功能模块
try:
    from features import (generate_next_month_template, export_to_pdf,
                          generate_person_cards, compare_months, backup_output,
                          data_preview, generate_ranking_html,
                          generate_project_management_html, smart_episode_assignment,
                          generate_person_trend_html, validate_project_data,
                          list_backups, cleanup_backups,
                          advanced_filter, correct_record,
                          generate_project_template,
                          create_config_snapshot, list_config_snapshots,
                          restore_config_snapshot, start_web_server,
                          validate_episode_assignments)
    HAS_FEATURES = True
except ImportError as e:
    HAS_FEATURES = False
    print(f'features import error: {e}')

# 优先使用 Python 3.13（保证有 pandas/openpyxl）
_PY313 = r'C:\Users\Admin\AppData\Local\Programs\Python\Python313\python.exe'
if os.path.exists(_PY313):
    PYTHON_EXE = _PY313
else:
    PYTHON_EXE = sys.executable

ROLES = ['一卡剪辑', '二卡剪辑', '剪辑助理', '剪辑组长', '小组长']
ROLE_ICONS = {'一卡剪辑': '🟢', '二卡剪辑': '🔵', '剪辑助理': '🟣', '剪辑组长': '🟠', '小组长': '🟡'}

# ============ 配色方案 v7.0 - 清晰工作台 ============
C = {
    # 全局背景
    'bg':            '#f4f7f8',
    'card':          '#ffffff',
    'card_hover':    '#f6fbfa',
    'card_border':   '#dce6e5',
    # 主色调
    'accent':        '#007c70',
    'accent_h':      '#00675e',
    'accent_a':      '#00554e',
    'accent_l':      '#e3f4f1',
    # 功能色
    'green':         '#0ea16b',
    'green_l':       '#e6f7f0',
    'blue':          '#2b7fff',
    'blue_l':        '#eaf2ff',
    'purple':        '#6d5bd0',
    'purple_l':      '#f0effb',
    'orange':        '#f0641a',
    'orange_l':      '#fff3eb',
    'red':           '#e53e3e',
    'red_l':         '#fef0f0',
    'teal':          '#0e9388',
    'teal_l':        '#e6f7f5',
    'pink':          '#db2777',
    'amber':         '#d97706',
    'cyan':          '#0e8fa6',
    'indigo':        '#536acb',
    'gray':          '#6b7280',
    'slate':         '#475569',
    # 文字
    'text':          '#111827',
    'text2':         '#4b5563',
    'text3':         '#9ca3af',
    'placeholder':   '#d1d5db',
    # 边框
    'border':        '#e5e7eb',
    'border_l':      '#f3f4f6',
    # 头部 (深色)
    'hdr_bg':        '#102827',
    'hdr_text':      '#f2f8f7',
    'hdr_sub':       '#8ed9d0',
    # 日志终端
    'log_bg':        '#0d1117',
    'log_fg':        '#c9d1d9',
    # 状态栏
    'status_bg':     '#f9fafb',
    'status_fg':     '#9ca3af',
}

# 角色配色 (前景色, 背景色)
ROLE_COLORS = {
    '一卡剪辑':  ('#0ea16b', '#e6f7f0'),
    '二卡剪辑':  ('#2b7fff', '#eaf2ff'),
    '剪辑助理':  ('#7c3aed', '#f3f0ff'),
    '剪辑组长':  ('#f0641a', '#fff3eb'),
    '小组长':    ('#d97706', '#fef6ec'),
}

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("提成表生成 · AI后期剪辑组")
        self.root.geometry("1180x800")
        self.root.minsize(960, 680)
        self.root.configure(bg=C['bg'])

        self._setup_ttk_style()

        try: self.root.iconbitmap(os.path.join(SCRIPT_DIR, 'icon.ico'))
        except: pass

        self.cfg = self._load_config()
        self.project_file = os.path.join(SCRIPT_DIR, '一组AI项目.xlsx')
        self.template_file = os.path.join(SCRIPT_DIR, 'AI后期剪辑提成一组最新.xlsx')
        self.output_dir = SCRIPT_DIR
        if not os.path.exists(self.template_file):
            self.template_file = os.path.join(SCRIPT_DIR, 'AI后期剪辑提成一组模板.xlsx')
        self.auto_backup = tk.BooleanVar(value=True)
        self._current_overtime_map = {}
        self._generation_year = None
        self._watcher = False
        self._web_server_instance = None
        self.build_ui()

        self.root.bind('<Control-g>', lambda e: self.run())
        self.root.bind('<Control-o>', lambda e: os.startfile(self.output_dir))
        self.root.bind('<Control-r>', lambda e: self.open_role_editor())
        self.root.bind('<Control-G>', lambda e: self.run())
        self.root.bind('<Control-O>', lambda e: os.startfile(self.output_dir))
        self.root.bind('<Control-R>', lambda e: self.open_role_editor())
        self.check_files()

    def _setup_ttk_style(self):
        s = ttk.Style()
        s.theme_use('clam')

        s.configure('TNotebook', background=C['bg'], borderwidth=0, tabmargins=(0, 0, 0, 8))
        s.configure('TNotebook.Tab',
            font=('Microsoft YaHei', 10),
            padding=(20, 9),
            background=C['bg'],
            foreground=C['text2'],
            borderwidth=0)
        s.map('TNotebook.Tab',
            background=[('selected', C['accent_l'])],
            foreground=[('selected', C['accent'])],
            expand=[('selected', [0, 0, 0, 0])])

        s.configure('TProgressbar', thickness=3,
            background=C['accent'], troughcolor=C['border_l'])

        s.configure('Treeview', font=('Microsoft YaHei', 9),
            rowheight=30, background=C['card'],
            fieldbackground=C['card'], foreground=C['text'])
        s.configure('Treeview.Heading', font=('Microsoft YaHei', 9, 'bold'),
            background='#edf3f2', foreground=C['text'], relief='flat')
        s.map('Treeview',
            background=[('selected', C['accent_l'])],
            foreground=[('selected', C['text'])])

        s.configure('TCombobox',
            fieldbackground=C['card'], background=C['card'],
            arrowcolor=C['text'])

    # ============ 通用组件 ============

    @staticmethod
    def _card(parent, **pack):
        """白色卡片，统一圆角风格（通过 highlightthickness 模拟边框）"""
        f = tk.Frame(parent, bg=C['card'], highlightthickness=1,
                     highlightbackground=C['card_border'])
        if pack: f.pack(**pack)
        return f

    @staticmethod
    def _hdr(parent, icon, title):
        """区块标题：图标 + 文字"""
        bar = tk.Frame(parent, bg=C['card'])
        bar.pack(fill='x', padx=14, pady=(12, 6))
        tk.Label(bar, text=icon, font=('Microsoft YaHei', 13), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=(0, 6))
        tk.Label(bar, text=title, font=('Microsoft YaHei', 12, 'bold'),
                 bg=C['card'], fg=C['text']).pack(side='left')
        sep = tk.Frame(parent, bg=C['border_l'], height=1)
        sep.pack(fill='x', padx=14, pady=(0, 6))

    @staticmethod
    def _btn(parent, text, color, cmd, font_size=10, padx=16, pady=8):
        """统一命令按钮"""
        hc = C['accent_a'] if color == C['accent'] else color
        btn = tk.Label(parent, text=text, font=('Microsoft YaHei', font_size, 'bold'),
                       bg=color, fg='white', padx=padx, pady=pady,
                       cursor='hand2', relief='flat')
        btn.bind('<Button-1>', lambda e: cmd())
        btn.bind('<Enter>', lambda e: btn.configure(bg=C['accent_h'] if color == C['accent'] else hc))
        btn.bind('<Leave>', lambda e: btn.configure(bg=color))
        return btn

    @staticmethod
    def _tooltip(widget, text):
        """为图标按钮提供鼠标悬停说明。"""
        tip = [None]

        def show(_event):
            if tip[0] is not None:
                return
            popup = tk.Toplevel(widget)
            popup.wm_overrideredirect(True)
            popup.configure(bg='#172120')
            x = widget.winfo_rootx()
            y = widget.winfo_rooty() + widget.winfo_height() + 6
            popup.geometry(f'+{x}+{y}')
            tk.Label(popup, text=text, bg='#172120', fg='white',
                     font=('Microsoft YaHei', 8), padx=8, pady=4).pack()
            tip[0] = popup

        def hide(_event):
            if tip[0] is not None:
                tip[0].destroy()
                tip[0] = None

        widget.bind('<Enter>', show, add='+')
        widget.bind('<Leave>', hide, add='+')

    def _load_config(self):
        from config_loader import load_config
        return load_config(CONFIG_PATH).to_dict()

    def _save_config(self):
        from config_loader import save_config
        from models import AppConfig
        save_config(AppConfig.from_dict(self.cfg), CONFIG_PATH)

    def _load_gc_module(self):
        """加载 generate_commission 模块并注入当前配置"""
        import sys as _sys
        _sys.path.insert(0, SCRIPT_DIR)
        import generate_commission as gc
        gc.set_config(self.cfg)
        return gc, _sys

    # ============ UI 构建 ============

    def build_ui(self):
        # ---- 应用栏 ----
        hdr = tk.Frame(self.root, bg=C['hdr_bg'], height=72)
        hdr.pack(fill='x'); hdr.pack_propagate(False)
        hl = tk.Frame(hdr, bg=C['hdr_bg'])
        hl.pack(side='left', padx=24, pady=12)
        tk.Label(hl, text='AI 后期剪辑', font=('Microsoft YaHei', 15, 'bold'),
                 fg=C['hdr_text'], bg=C['hdr_bg']).pack(anchor='w')
        tk.Label(hl, text='提成工作台  ·  月度核算与报表生成',
                 font=('Microsoft YaHei', 8), fg=C['hdr_sub'],
                 bg=C['hdr_bg']).pack(anchor='w', pady=(1, 0))

        actions = tk.Frame(hdr, bg=C['hdr_bg'])
        actions.pack(side='right', padx=20)
        open_btn = self._btn(actions, '📂', '#234846',
                             lambda: os.startfile(self.output_dir), padx=12, pady=7)
        open_btn.pack(side='left', padx=3)
        self._tooltip(open_btn, '打开输出目录')
        role_btn = self._btn(actions, '⚙', '#234846', self.open_role_editor, padx=12, pady=7)
        role_btn.pack(side='left', padx=3)
        self._tooltip(role_btn, '管理人员角色')
        tk.Label(actions, text='v7.0', font=('Consolas', 9), fg='#89aaa7',
                 bg=C['hdr_bg']).pack(side='left', padx=(12, 0))

        # ---- 主导航 ----
        nb = ttk.Notebook(self.root)
        nb.pack(fill='both', expand=True, padx=22, pady=(14, 0))

        tab1 = tk.Frame(nb, bg=C['bg']); nb.add(tab1, text='  工作台  ')
        self._build_tab_main(tab1)
        tab2 = tk.Frame(nb, bg=C['bg']); nb.add(tab2, text='  工具箱  ')
        self._build_tab_tools(tab2)
        tab3 = tk.Frame(nb, bg=C['bg']); nb.add(tab3, text='  分析与管理  ')
        self._build_tab_advanced(tab3)

        # ---- 状态栏 ----
        bar = tk.Frame(self.root, bg=C['status_bg'], height=28)
        bar.pack(fill='x', side='bottom'); bar.pack_propagate(False)
        self.st = tk.Label(bar, text='  ●  就绪 · 配置已加载', font=('Microsoft YaHei', 9),
                           bg=C['status_bg'], fg=C['status_fg'],
                           anchor='w', padx=14)
        self.st.pack(fill='x')

    def _build_tab_main(self, p):
        """主工作台：准备输入、发起生成、查看执行结果。"""
        intro = tk.Frame(p, bg=C['bg'])
        intro.pack(fill='x', pady=(2, 12))
        tk.Label(intro, text='月度提成生成', font=('Microsoft YaHei', 18, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(anchor='w')
        tk.Label(intro, text='核对文件与角色后，生成 Excel、统计简报、仪表盘和个人绩效卡片。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack(anchor='w', pady=(3, 0))

        c1 = self._card(p, fill='x', pady=(0, 12))
        header = tk.Frame(c1, bg=C['card']); header.pack(fill='x', padx=16, pady=(12, 5))
        tk.Label(header, text='输入与输出', font=('Microsoft YaHei', 11, 'bold'),
                 bg=C['card'], fg=C['text']).pack(side='left')
        tk.Label(header, text='生成前可直接更换文件', font=('Microsoft YaHei', 8),
                 bg=C['card'], fg=C['text3']).pack(side='right')
        file_grid = tk.Frame(c1, bg=C['card']); file_grid.pack(fill='x', padx=12, pady=(0, 12))
        for col, (label, attr, cmd) in enumerate([
            ('项目数据', 'pf_label', self._select_project),
            ('提成模板', 'tf_label', self._select_template),
            ('输出目录', 'od_label', self._select_output_dir)]):
            cell = tk.Frame(file_grid, bg='#f8fbfa', highlightthickness=1,
                            highlightbackground=C['card_border'])
            cell.grid(row=0, column=col, sticky='nsew', padx=4)
            tk.Label(cell, text=label, font=('Microsoft YaHei', 8, 'bold'),
                     bg='#f8fbfa', fg=C['text2']).pack(anchor='w', padx=10, pady=(8, 1))
            lbl = tk.Label(cell, text='', font=('Microsoft YaHei', 8), bg='#f8fbfa',
                           fg=C['text3'], anchor='w')
            lbl.pack(fill='x', padx=10, pady=(0, 7))
            setattr(self, attr, lbl)
            change = tk.Label(cell, text='更换', font=('Microsoft YaHei', 8, 'bold'),
                              bg=C['accent_l'], fg=C['accent'], cursor='hand2', padx=9, pady=2)
            change.pack(anchor='e', padx=8, pady=(0, 8))
            change.bind('<Button-1>', lambda e, c=cmd: c())
            file_grid.grid_columnconfigure(col, weight=1, uniform='file')
        self._refresh_file_labels()

        body = tk.Frame(p, bg=C['bg']); body.pack(fill='both', expand=True)
        left = tk.Frame(body, bg=C['bg'], width=385)
        left.pack(side='left', fill='both', padx=(0, 12)); left.pack_propagate(False)
        right = tk.Frame(body, bg=C['bg'])
        right.pack(side='left', fill='both', expand=True)

        cc = self._card(left, fill='both', expand=True, pady=(0, 8))
        self._hdr(cc, '👥', '本组角色配置')
        self.role_tags = tk.Frame(cc, bg=C['card'])
        self.role_tags.pack(fill='both', expand=True, padx=10, pady=(0, 8))
        self._refresh_role_tags()

        run_card = self._card(left, fill='x')
        tk.Label(run_card, text='准备完成后开始生成', font=('Microsoft YaHei', 9, 'bold'),
                 bg=C['card'], fg=C['text']).pack(anchor='w', padx=14, pady=(12, 2))
        tk.Label(run_card, text='将先标记超时集数，再生成全部报表。', font=('Microsoft YaHei', 8),
                 bg=C['card'], fg=C['text3']).pack(anchor='w', padx=14, pady=(0, 10))
        btn_row = tk.Frame(run_card, bg=C['card']); btn_row.pack(fill='x')
        self.run_btn = self._btn(btn_row, '▶  开始生成', C['accent'],
                                  self.run, font_size=12, padx=28, pady=10)
        self.run_btn.pack(side='left', fill='x', expand=True, padx=(14, 4), pady=(0, 12))
        role_small = self._btn(btn_row, '⚙', C['orange'], self.open_role_editor, padx=13, pady=10)
        role_small.pack(side='left', padx=4, pady=(0, 12))
        self._tooltip(role_small, '管理角色')
        self.progress = ttk.Progressbar(run_card, mode='indeterminate')
        self.progress.pack(fill='x', padx=14, pady=(0, 8))
        aux = tk.Frame(run_card, bg=C['card']); aux.pack(fill='x', padx=14, pady=(0, 10))
        tk.Checkbutton(aux, text='自动备份', variable=self.auto_backup,
                       font=('Microsoft YaHei', 8), bg=C['card'], fg=C['text3'],
                       selectcolor=C['card'], activebackground=C['card']).pack(side='left')
        tk.Label(aux, text='Ctrl+G 生成  ·  Ctrl+R 角色', font=('Microsoft YaHei', 7),
                 bg=C['card'], fg=C['text3']).pack(side='right')

        c3 = self._card(right, fill='both', expand=True)
        self._hdr(c3, '📝', '运行记录')
        tw = tk.Frame(c3, bg=C['log_bg'])
        tw.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        self.log_txt = tk.Text(tw, font=('Consolas', 9),
                               bg=C['log_bg'], fg=C['log_fg'],
                               insertbackground='#58a6ff',
                               relief='flat', padx=10, pady=8,
                               wrap='word', state='disabled',
                               selectbackground='#1f6feb')
        self.log_txt.pack(side='left', fill='both', expand=True)
        sb = tk.Scrollbar(tw, bg=C['log_bg'], troughcolor=C['log_bg'])
        sb.pack(side='right', fill='y')
        self.log_txt.configure(yscrollcommand=sb.set)
        sb.configure(command=self.log_txt.yview)

    def _build_tab_tools(self, p):
        canvas = tk.Canvas(p, bg=C['bg'], highlightthickness=0)
        sb = ttk.Scrollbar(p, orient='vertical', command=canvas.yview)
        sf = tk.Frame(canvas, bg=C['bg'])
        sf.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        wid = canvas.create_window((0,0), window=sf, anchor='nw')
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side='left', fill='both', expand=True, padx=(0, 6))
        sb.pack(side='right', fill='y')
        canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(int(-e.delta/120),'units'))
        def _fw(e): canvas.itemconfig(wid, width=e.width)
        canvas.bind('<Configure>', _fw)

        # Use Unicode text directly instead of escape sequences
        groups = [
            ('📊 数据工具', [
                ('📋','数据预览', C['teal'], self._preview_data, '预览解析后的人员集数、绩效和提成汇总'),
                ('✅','数据校验', C['green'], self._validate_data, '校验项目数据日期、人名、重复分配等问题'),
                ('🔍','项目去重', C['red'], self._check_duplicates, '检查相同项目ID是否对应了不同项目名称'),
                ('📐','智能分集', C['pink'], self._smart_assign, '按角色权重自动分配各人负责集数'),
            ]),
            ('📈 报表生成', [
                ('📊','月份对比', C['orange'], self._compare_months, '选择两个月提成表对比集数和提成变化'),
                ('🏆','组内排名', C['amber'], self._gen_ranking, '生成月度集数排行和提成排行HTML'),
                ('🗂','项目管理', C['purple'], self._gen_project_mgmt, '生成项目清单视图按交付日期排序'),
                ('🃏','绩效卡片', C['blue'], self._gen_cards, '每人独立HTML绩效卡片含集数达标提成'),
            ]),
            ('🔧 辅助工具', [
                ('📅','下月模板', C['indigo'], self._gen_next_template, '基于当前模板自动创建下月空白模板'),
                ('📤','导出PDF', C['cyan'], self._export_pdf, '将Excel提成表导出为A3横版PDF'),
                ('🏷','提成规则', C['slate'], self._edit_rules, '可视化编辑各角色的基准集数单价'),
                ('📥','模板下载', C['teal'], self._download_template, '下载标准化项目数据录入模板Excel'),
            ]),
            ('⚡ 高级工具', [
                ('🔎','高级筛选', C['blue'], self._advanced_filter, '多条件组合筛选支持导出CSV'),
                ('✏','数据修正', '#f0641a', self._data_correction, '双击修正集数或删除重新生成提成表'),
                ('🌐','Web服务', C['accent'], self._web_server, '启动本地HTTP团队浏览器查看报表'),
                ('📸','配置快照', C['purple'], self._config_snapshot, '保存配置历史支持一键回滚'),
                ('🔄','文件监控', C['green'], self._toggle_watch, '检测数据变化自动提示重新生成'),
            ]),
        ]

        for gtitle, tools in groups:
            gf = tk.Frame(sf, bg=C['bg'])
            gf.pack(fill='x', padx=2, pady=(14, 7))
            tk.Label(gf, text=gtitle, font=('Microsoft YaHei', 11, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(side='left')
            tk.Frame(gf, bg=C['border'], height=1).pack(
                side='left', fill='x', expand=True, padx=(10,0))

            grid = tk.Frame(sf, bg=C['bg'])
            grid.pack(fill='x', padx=2)
            for i, (icon, name, color, cmd, desc) in enumerate(tools):
                card = tk.Frame(grid, bg=C['card'], highlightthickness=1,
                                highlightbackground=C['card_border'],
                                cursor='hand2')
                card.grid(row=i//3, column=i%3, padx=4, pady=4, sticky='nsew')
                tk.Label(card, text=icon, font=('Microsoft YaHei', 19),
                         bg=C['card']).pack(pady=(12, 0))
                tk.Label(card, text=name, font=('Microsoft YaHei', 10, 'bold'),
                         bg=C['card'], fg=C['text']).pack(pady=(4, 0))
                tk.Label(card, text=desc, font=('Microsoft YaHei', 7),
                         bg=C['card'], fg=C['text3'], wraplength=210,
                         justify='center').pack(padx=10, pady=(2, 12))
                # bottom accent bar
                tk.Frame(card, bg=color, height=3).pack(fill='x', side='bottom')
                for ch in list(card.children.values()):
                    ch.bind('<Button-1>', lambda e, c=cmd: c())
                card.bind('<Button-1>', lambda e, c=cmd: c())
                card.bind('<Enter>', lambda e, f=card: f.configure(bg=C['card_hover']))
                card.bind('<Leave>', lambda e, f=card: f.configure(bg=C['card']))
            for c in range(3): grid.grid_columnconfigure(c, weight=1)

    def _build_tab_advanced(self, p):
        # 个人趋势
        c1 = self._card(p, fill='x', padx=10, pady=(10,6))
        self._hdr(c1, '📈', '个人月度趋势')
        sf = tk.Frame(c1, bg=C['card']); sf.pack(fill='x', padx=14, pady=(0,10))
        self._trend_var = tk.StringVar()
        names = sorted(self.cfg.get('人员角色', {}).keys())
        if names:
            self._trend_var.set(names[0])
            ttk.Combobox(sf, textvariable=self._trend_var, values=names,
                         state='readonly', font=('Microsoft YaHei', 10),
                         width=10).pack(side='left', padx=(0,8))
        self._btn(sf, '生成趋势图', C['accent'], self._gen_trend).pack(side='left')

        # 备份管理
        c2 = self._card(p, fill='both', expand=True, padx=10, pady=(4,8))
        self._hdr(c2, '🗄', '备份管理')
        hf = tk.Frame(c2, bg=C['card']); hf.pack(fill='x', padx=14, pady=(0,4))
        self._btn(hf, '清理旧备份', C['red'], self._manage_backups,
                  font_size=9, padx=10, pady=4).pack(side='right')
        self._backup_list = tk.Text(c2, font=('Consolas', 9),
                                    bg=C['log_bg'], fg=C['log_fg'],
                                    relief='flat', padx=10, pady=8,
                                    height=10, wrap='word',
                                    selectbackground='#1f6feb')
        self._backup_list.pack(fill='both', expand=True, padx=10, pady=(0,10))
        self._refresh_backups()
    def _refresh_role_tags(self):
        for w in self.role_tags.winfo_children():
            w.destroy()
        if not self.cfg: return
        rm = self.cfg.get('人员角色', {})
        if not rm:
            tk.Label(self.role_tags, text='暂无人员，请点击「角色配置」添加',
                     font=('Microsoft YaHei', 9), bg=C['card'],
                     fg=C['text3']).pack(anchor='w', pady=10)
            return

        total = len(rm)
        tk.Label(self.role_tags, text=f'共 {total} 人',
                 font=('Microsoft YaHei', 8), bg=C['card'],
                 fg=C['text3']).pack(anchor='w', pady=2)

        for role in ROLES:
            names = [n for n, r in rm.items() if r == role]
            if not names:
                continue
            fg, bg_c = ROLE_COLORS.get(role, ('#666', '#f0f0f0'))
            tag = tk.Frame(self.role_tags, bg=bg_c)
            tag.pack(fill='x', pady=3, padx=2)
            tk.Label(tag, text=f'  {role}  ', font=('Microsoft YaHei', 9, 'bold'),
                     bg=bg_c, fg=fg, padx=6, pady=4).pack(side='left')
            sep = tk.Frame(tag, bg=fg, width=1)
            sep.pack(side='left', fill='y', padx=5, pady=5)
            tk.Label(tag, text='、'.join(names),
                     font=('Microsoft YaHei', 9),
                     bg=bg_c, fg=C['text2'], padx=2, pady=4).pack(side='left')

    # ============ 角色编辑器 ============

    def open_role_editor(self):
        if not self.cfg:
            messagebox.showwarning('错误', '无法加载 config.json'); return
        dlg = tk.Toplevel(self.root)
        dlg.title('角色配置编辑器'); dlg.geometry('620x620')
        dlg.minsize(500, 500); dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='👥 人员角色配置', font=('Microsoft YaHei', 16, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=16)
        tk.Label(dlg, text='修改角色 / 添加人员 / 删除人员 · 保存后即时生效',
                 font=('Microsoft YaHei', 9), bg=C['bg'],
                 fg=C['text3']).pack(pady=0)

        # ---- 新增人员栏 ----
        add_bar = tk.Frame(dlg, bg=C['bg'])
        add_bar.pack(fill='x', padx=20, pady=0)
        tk.Label(add_bar, text='新增:', font=('Microsoft YaHei', 10),
                 bg=C['bg'], fg=C['text']).pack(side='left', padx=0)
        add_name = tk.Entry(add_bar, font=('Microsoft YaHei', 10), width=12)
        add_name.pack(side='left', padx=0)
        add_role_var = tk.StringVar(value='一卡剪辑')
        ttk.Combobox(add_bar, textvariable=add_role_var, values=ROLES,
                     state='readonly', font=('Microsoft YaHei', 10), width=10).pack(side='left', padx=0)
        rm_ref = self.cfg['人员角色']
        order_ref = self.cfg.get('人员排序', list(rm_ref.keys()))

        # 刷新函数——重建整个列表
        def rebuild_person_list():
            for w in sf.winfo_children():
                w.destroy()
            for i, name in enumerate(order_ref):
                if name not in rm_ref:
                    continue
                row = tk.Frame(sf, bg='white', highlightthickness=1,
                               highlightbackground=C['border'])
                row.grid(row=i, column=0, padx=5, pady=2, sticky='ew')
                # 删除按钮
                btn_del = tk.Button(row, text='✕', font=('Microsoft YaHei', 9, 'bold'),
                                    bg='#fee2e2', fg='#dc2626', relief='flat',
                                    cursor='hand2', padx=6, pady=3,
                                    activebackground='#fecaca',
                                    command=lambda n=name: _delete_person(n))
                btn_del.pack(side='left', padx=6, pady=5)
                # 姓名
                tk.Label(row, text=name, font=('Microsoft YaHei', 10, 'bold'),
                         bg='white', fg=C['text'], width=8, anchor='w').pack(side='left', padx=2, pady=5)
                # 角色下拉
                current = rm_ref.get(name, '一卡剪辑')
                var = tk.StringVar(value=current)
                role_vars[name] = var
                cb = ttk.Combobox(row, textvariable=var, values=ROLES,
                                  state='readonly', font=('Microsoft YaHei', 10), width=10)
                cb.pack(side='left', padx=0, pady=5)
                # 颜色点
                color_map = {'一卡剪辑': '#27ae60', '二卡剪辑': '#3b82f6',
                             '剪辑助理': '#8b5cf6', '剪辑组长': '#f59e0b'}
                dot = tk.Label(row, text='●', font=('Microsoft YaHei', 14),
                               fg=color_map.get(current, '#333'), bg='white')
                dot.pack(side='left', pady=5)
                def _on_change(*_, lbl=dot, v=var):
                    lbl.configure(fg=color_map.get(v.get(), '#333'))
                var.trace_add('write', _on_change)
            sf.grid_columnconfigure(0, weight=1)

        def _delete_person(name):
            if not messagebox.askyesno('确认删除', f'确定要删除人员 "{name}" 吗？\n\n此操作不可恢复。'):
                return
            rm_ref.pop(name, None)
            if name in order_ref:
                order_ref.remove(name)
            role_vars.pop(name, None)
            self.cfg['人员排序'] = order_ref
            rebuild_person_list()

        def _add_person():
            name = add_name.get().strip()
            if not name:
                messagebox.showwarning('提示', '请输入人员姓名'); return
            if name in rm_ref:
                messagebox.showwarning('提示', f'人员 "{name}" 已存在'); return
            rm_ref[name] = add_role_var.get()
            order_ref.append(name)
            self.cfg['人员排序'] = order_ref
            add_name.delete(0, 'end')
            rebuild_person_list()
            # 滚到底部
            canvas.yview_moveto(1.0)

        tk.Button(add_bar, text='＋ 添加', font=('Microsoft YaHei', 10, 'bold'),
                  bg=C['blue'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=3, command=_add_person).pack(side='left', padx=4)
        # 绑定回车
        add_name.bind('<Return>', lambda e: _add_person())

        # ---- 可滚动人员列表 ----
        canvas = tk.Canvas(dlg, bg=C['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(dlg, orient='vertical', command=canvas.yview)
        sf = tk.Frame(canvas, bg=C['bg'])
        sf.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((10, 0), window=sf, anchor='nw', width=570)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True, padx=20, pady=0)
        scrollbar.pack(side='right', fill='y', pady=0)

        role_vars = {}
        rebuild_person_list()

        # ---- 底部按钮 ----
        bf = tk.Frame(dlg, bg=C['bg'])
        bf.pack(fill='x', padx=20, pady=0)

        def save():
            for n, v in role_vars.items():
                if n in rm_ref:
                    rm_ref[n] = v.get()
            self.cfg['人员角色'] = rm_ref
            self.cfg['人员排序'] = order_ref

            # 同步更新小组中的成员列表
            all_names = set(rm_ref.keys())
            groups = self.cfg.get('小组', {})
            for gname, ginfo in groups.items():
                members = ginfo.get('成员', [])
                new_members = [m for m in members if m in all_names]
                ginfo['成员'] = new_members
                leader = ginfo.get('组长', '')
                if leader not in all_names:
                    ginfo['组长'] = new_members[0] if new_members else ''

            self._save_config()
            self._refresh_role_tags()
            self._log('✅ 角色配置已更新并保存')
            dlg.destroy()
            messagebox.showinfo('保存成功', '角色配置已保存！\n\n下次生成时生效。')

        tk.Button(bf, text='💾 保存配置', font=('Microsoft YaHei', 12, 'bold'),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=30, pady=8, command=save).pack(side='right', padx=10)
        tk.Button(bf, text='取消', font=('Microsoft YaHei', 11),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=24, pady=8, command=dlg.destroy).pack(side='right')
        canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(int(-e.delta/120), 'units'))
        dlg.protocol('WM_DELETE_WINDOW', lambda: [canvas.unbind_all('<MouseWheel>'), dlg.destroy()])

    # ============ 文件选择 ============

    def _select_project(self):
        path = filedialog.askopenfilename(
            title='选择项目数据文件',
            filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')],
            initialdir=SCRIPT_DIR)
        if path:
            self.project_file = path
            self._refresh_file_labels()
            self.check_files()

    def _select_template(self):
        path = filedialog.askopenfilename(
            title='选择模板文件',
            filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')],
            initialdir=SCRIPT_DIR)
        if path:
            self.template_file = path
            self._refresh_file_labels()
            self.check_files()

    def _select_output_dir(self):
        path = filedialog.askdirectory(
            title='选择输出目录',
            initialdir=self.output_dir)
        if path:
            self.output_dir = path
            self._refresh_file_labels()
            self.check_files()

    def _refresh_file_labels(self):
        def _short(fpath):
            if not fpath: return '（未选择）'
            name = os.path.basename(fpath)
            exists = os.path.exists(fpath)
            icon = '✅' if exists else '❌'
            return f'{icon} {name}'
        self.pf_label.configure(text=_short(self.project_file))
        self.tf_label.configure(text=_short(self.template_file))
        self.od_label.configure(text=f'📁 {self.output_dir}')

    # ============ 功能：生成下月模板 ============
    def _gen_next_template(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return
        try:
            self._log('📅 正在生成下月模板...')
            self._log(f'   源模板: {os.path.basename(self.template_file)}')
            path, msg = generate_next_month_template(self.template_file, self.output_dir)
            if path:
                self._log(f'✅ {msg}')
                self._log(f'   文件: {os.path.basename(path)}')
                try: os.startfile(path)
                except: pass
                messagebox.showinfo('完成', f'{msg}\n\n文件已自动打开。')
            else:
                self._log(f'❌ {msg}')
                messagebox.showerror('失败', msg)
        except Exception as e:
            self._log(f'❌ 生成下月模板失败: {e}')
            messagebox.showerror('失败', f'生成失败:\n{e}')

    # ============ 功能：导出PDF ============
    def _export_pdf(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return
        path = filedialog.askopenfilename(
            title='选择要导出的Excel提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=SCRIPT_DIR)
        if not path:
            return
        self._log(f'📤 正在导出PDF: {os.path.basename(path)}')
        try:
            pdf_path = export_to_pdf(path)
            self._log(f'✅ PDF已生成: {os.path.basename(pdf_path)}')
            try: os.startfile(pdf_path)
            except: pass
            messagebox.showinfo('完成', f'PDF导出成功！\n\n📄 {os.path.basename(pdf_path)}')
        except Exception as e:
            self._log(f'❌ PDF导出失败: {e}')
            messagebox.showerror('失败', f'PDF导出失败:\n{e}')

    # ============ 功能：个人绩效卡片 ============
    def _gen_cards(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return
        self._log('🃏 正在生成个人绩效卡片...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            if not records:
                self._log('⚠️ 无数据，请检查项目文件')
                _sys.path.pop(0)
                return
            cd = gc.compute_commission(records, group_pids)
            card_paths = generate_person_cards(records, cd, CARDS_DIR)
            _sys.path.pop(0)

            if card_paths:
                self._log(f'✅ 已生成 {len(card_paths)-1} 张个人绩效卡片 -> 个人绩效卡片/')
                try: os.startfile(card_paths[0])
                except: pass
            else:
                self._log('⚠️ 未能生成卡片')
        except Exception as e:
            self._log(f'❌ 卡片生成失败: {e}')

    # ============ 功能：多月份对比 ============
    def _compare_months(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return
        f1 = filedialog.askopenfilename(
            title='选择第一个月份的提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=SCRIPT_DIR)
        if not f1: return
        f2 = filedialog.askopenfilename(
            title='选择第二个月份的提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=os.path.dirname(f1))
        if not f2: return
        self._log(f'📊 正在对比两个月份的提成表...')
        try:
            label1, label2, diffs = compare_months(f1, f2)
            if not diffs:
                self._log(f'✅ 两个月份数据完全一致，无差异。')
                messagebox.showinfo('对比结果', f'{label1} ↔ {label2}\n\n数据完全一致，无差异。')
                return
            self._log(f'{label1} ↔ {label2} 对比: {len(diffs)}人有变化')
            msg = f'{label1}  →  {label2}\n{"="*60}\n'
            msg += f'{"姓名":　<6s} {"上月集数":>6s} {"本月集数":>6s} {"集数变化":>6s} {"上月提成":>8s} {"本月提成":>8s} {"提成变化":>8s}\n'
            for nm, e1, e2, de, c1, c2, dc in diffs:
                de_str = f'+{de}' if de > 0 else str(de)
                dc_str = f'+{dc}' if dc > 0 else str(dc)
                msg += f'{nm:　<6s} {e1:>6d} {e2:>6d} {de_str:>6s} {c1:>8d} {c2:>8d} {dc_str:>8s}\n'
            self._log(msg)
            # 弹窗显示
            dlg = tk.Toplevel(self.root)
            dlg.title('多月份对比')
            dlg.geometry('700x500')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text=f'{label1}  ↔  {label2}', font=('Microsoft YaHei', 14, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=10)
            txt = tk.Text(dlg, font=('Consolas', 10), bg=C['log_bg'], fg=C['log_fg'],
                          relief='flat', padx=12, pady=10)
            txt.insert('1.0', msg)
            txt.configure(state='disabled')
            txt.pack(fill='both', expand=True, padx=20, pady=0)
        except Exception as e:
            self._log(f'❌ 对比失败: {e}')
            messagebox.showerror('失败', f'对比失败:\n{e}')

    # ============ 功能：提成规则面板 ============
    def _edit_rules(self):
        dlg = tk.Toplevel(self.root)
        dlg.title('提成规则设置')
        dlg.geometry('500x420')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='🏷️ 提成规则配置', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=20)
        tk.Label(dlg, text='修改后点击保存，下次生成生效', font=('Microsoft YaHei', 9),
                 bg=C['bg'], fg=C['text3']).pack()

        rules = self.cfg.get('rules', {})
        vars_map = {}
        for role_name in ['一卡剪辑', '二卡剪辑', '剪辑助理']:
            frm = tk.Frame(dlg, bg=C['card'], highlightthickness=1,
                           highlightbackground=C['border'])
            frm.pack(fill='x', padx=20, pady=4)
            tk.Label(frm, text=f'{role_name}', font=('Microsoft YaHei', 11, 'bold'),
                     bg=C['card'], fg=C['text'], width=12, anchor='w').pack(side='left', padx=10, pady=8)

            r = rules.get(role_name, {})
            for key, label in [('基准集数', '基准'), ('超额每集', '超额/集'), ('缺集每集扣', '缺扣/集')]:
                v = tk.IntVar(value=r.get(key, 70 if role_name == '一卡剪辑' else 120))
                vars_map[f'{role_name}|{key}'] = v
                sub = tk.Frame(frm, bg=C['card'])
                sub.pack(side='left', padx=4, pady=4)
                tk.Label(sub, text=label, font=('Microsoft YaHei', 8), bg=C['card'],
                         fg=C['text3']).pack()
                s = tk.Scale(sub, from_=0, to=300 if '集数' in key else 100,
                             orient='horizontal', variable=v, length=70,
                             bg=C['card'], fg=C['text'], highlightthickness=0)
                s.pack()

        # 组长规则
        frm_l = tk.Frame(dlg, bg=C['card'], highlightthickness=1,
                         highlightbackground=C['border'])
        frm_l.pack(fill='x', padx=20, pady=4)
        tk.Label(frm_l, text='剪辑组长', font=('Microsoft YaHei', 11, 'bold'),
                 bg=C['card'], fg=C['text'], width=12, anchor='w').pack(side='left', padx=10, pady=8)
        r_l = rules.get('剪辑组长', {})
        for key, label in [('每集单价', '单价/集'), ('组内每部提成', '每部奖')]:
            v = tk.IntVar(value=r_l.get(key, 20 if '单价' in key else 100))
            vars_map[f'剪辑组长|{key}'] = v
            sub = tk.Frame(frm_l, bg=C['card'])
            sub.pack(side='left', padx=4, pady=4)
            tk.Label(sub, text=label, font=('Microsoft YaHei', 8), bg=C['card'],
                     fg=C['text3']).pack()
            s = tk.Scale(sub, from_=0, to=300, orient='horizontal', variable=v,
                         length=70, bg=C['card'], fg=C['text'], highlightthickness=0)
            s.pack()

        def save_rules():
            for key, v in vars_map.items():
                role, field = key.split('|', 1)
                rules[role][field] = v.get()
            self._save_config()
            self._log('✅ 提成规则已更新并保存')
            dlg.destroy()
            messagebox.showinfo('保存成功', '提成规则已保存！\n\n下次生成时生效。')

        tk.Button(dlg, text='💾 保存规则', font=('Microsoft YaHei', 12, 'bold'),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=30, pady=8, command=save_rules).pack(pady=16)

    # ============ 功能：数据预览 ============
    def _preview_data(self):
        self._log('📋 正在加载数据预览...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            self._log(f'   项目文件: {os.path.basename(self.project_file)}')
            self._log(f'   人员数量: {len(gc.ALL_NAMES)}')
            df = pd.read_excel(self.project_file, header=None)
            self._log(f'   Excel行数: {len(df)}')
            records, group_pids = gc.parse_projects(df)
            self._log(f'   解析记录: {len(records)}')
            cd = gc.compute_commission(records, group_pids)
            preview = data_preview(records, cd)
            self._log(f'   预览行数: {len(preview)}')
            _sys.path.pop(0)

            if not preview:
                self._log('⚠️ 无数据可预览（可能项目文件名不在config人员名单中）')
                messagebox.showwarning('无数据', '解析到0条记录。\n\n请确认项目数据文件中的人员姓名与角色配置中的一致。')
                return

            # 弹窗展示
            dlg = tk.Toplevel(self.root)
            dlg.title('数据预览')
            dlg.geometry('750x520')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text='📋 数据预览', font=('Microsoft YaHei', 15, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=12)

            # Treeview
            cols = ('姓名', '角色', '集数', '项目数', '基准', '绩效', '提成')
            tree = ttk.Treeview(dlg, columns=cols, show='headings', height=18)
            widths = [80, 90, 60, 60, 60, 60, 80]
            for c, w in zip(cols, widths):
                tree.heading(c, text=c)
                tree.column(c, width=w, anchor='center')
            tree.pack(fill='both', expand=True, padx=20, pady=0)

            # 颜色标签
            tree.tag_configure('ok', foreground='#16a34a')
            tree.tag_configure('fail', foreground='#dc2626')
            for p in preview:
                tag = 'ok' if p['status'] == '是' else 'fail'
                tree.insert('', 'end', values=(
                    p['name'], p['role'], p['episodes'], p['projects'],
                    f'{p["quota"]}集' if p['quota'] > 0 else '无', p['status'],
                    f'{p["commission"]:,}'), tags=(tag,))
        except Exception as e:
            self._log(f'❌ 预览失败: {e}')
            import traceback; self._log(traceback.format_exc())

    # ============ 功能：组内排名 ============
    def _gen_ranking(self):
        self._log('🏆 正在生成组内排名...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            cd = gc.compute_commission(records, group_pids)
            _sys.path.pop(0)

            path = os.path.join(self.output_dir, '组内排名.html')
            generate_ranking_html(cd, path)
            self._log(f'✅ 组内排名已生成: {os.path.basename(path)}')
            try: os.startfile(path)
            except: pass
        except Exception as e:
            self._log(f'❌ 排名生成失败: {e}')

    # ============ 功能：项目管理视图 ============
    def _gen_project_mgmt(self):
        self._log('🗂️ 正在生成项目管理视图...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, _ = gc.parse_projects(df)
            _sys.path.pop(0)

            path = os.path.join(self.output_dir, '项目管理_项目清单.html')
            generate_project_management_html(records, path)
            self._log(f'✅ 项目管理视图已生成: {os.path.basename(path)}')
            try: os.startfile(path)
            except: pass
        except Exception as e:
            self._log(f'❌ 项目管理视图生成失败: {e}')

    # ============ 功能：文件监控 ============
    def _toggle_watch(self):
        if hasattr(self, '_watcher') and self._watcher:
            self._stop_watch()
            return
        self._watcher = True
        self._watch_mtime = os.path.getmtime(self.project_file)
        self.btn_watch.configure(text='🔄 监控中...', bg='#059669')
        self._log('🔄 文件监控已开启，检测到项目文件变化将自动提示...')
        self._watch_loop()

    def _watch_loop(self):
        if not getattr(self, '_watcher', False):
            return
        try:
            current_mtime = os.path.getmtime(self.project_file)
            if current_mtime != self._watch_mtime:
                self._watch_mtime = current_mtime
                self._log('🔔 检测到项目数据文件已更新！')
                if messagebox.askyesno('文件更新', '项目数据文件已更新，是否立即重新生成？'):
                    self.run()
        except Exception:
            pass
        if getattr(self, '_watcher', False):
            self.root.after(5000, self._watch_loop)

    def _stop_watch(self):
        self._watcher = False
        self.btn_watch.configure(text='🔄 开启监控', bg='#059669')
        self._log('🔴 文件监控已关闭')

    # ============ 功能：智能分集 ============
    def _smart_assign(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('📐 智能分集')
        dlg.geometry('750x680')
        dlg.minsize(650, 550)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        # 标题
        tk.Label(dlg, text='📐 智能分集工具', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=10)
        tk.Label(dlg, text='输入项目信息，选择剪辑人员，自动按角色区间分集',
                 font=('Microsoft YaHei', 8), bg=C['bg'], fg=C['text3']).pack()

        # 项目信息卡片
        c1 = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c1.pack(fill='x', padx=20, pady=6)
        in1 = tk.Frame(c1, bg=C['card']); in1.pack(fill='x', padx=12, pady=6)

        tk.Label(in1, text='项目名称:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).grid(row=0, column=0, sticky='w', pady=2)
        name_var = tk.StringVar()
        tk.Entry(in1, textvariable=name_var, font=('Microsoft YaHei', 10), width=40,
                 relief='solid', borderwidth=1).grid(row=0, column=1, padx=6, pady=2)

        tk.Label(in1, text='总集数:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).grid(row=1, column=0, sticky='w', pady=2)
        eps_var = tk.IntVar(value=100)
        tk.Spinbox(in1, from_=1, to=2000, textvariable=eps_var, font=('Microsoft YaHei', 10),
                   width=8, relief='solid', borderwidth=1).grid(row=1, column=1, padx=6, pady=2, sticky='w')

        tk.Label(in1, text='一卡区间(前N集):', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).grid(row=2, column=0, sticky='w', pady=2)
        range_var = tk.IntVar(value=15)
        tk.Spinbox(in1, from_=1, to=500, textvariable=range_var, font=('Microsoft YaHei', 10),
                   width=8, relief='solid', borderwidth=1).grid(row=2, column=1, padx=6, pady=2, sticky='w')

        # 人员选择卡片
        c2 = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c2.pack(fill='x', padx=20, pady=2)
        in2 = tk.Frame(c2, bg=C['card']); in2.pack(fill='x', padx=12, pady=6)
        tk.Label(in2, text='选择剪辑人员:', font=('Microsoft YaHei', 9, 'bold'),
                 bg=C['card'], fg=C['text']).pack(anchor='w')

        cb_frame = tk.Frame(in2, bg=C['card'])
        cb_frame.pack(fill='x', pady=3)

        check_vars = {}
        roles_map = self.cfg.get('人员角色', {})
        # 按角色排序展示
        role_order = {'小组长': 0, '一卡剪辑': 1, '二卡剪辑': 2, '剪辑助理': 3, '剪辑组长': 4}
        sorted_people = sorted(roles_map.keys(), key=lambda n: role_order.get(roles_map[n], 99))

        # 分组显示
        role_labels = [
            ('小组长', '🟡', lambda r: '小组长' in r),
            ('一卡剪辑', '🟢', lambda r: '一卡' in r and '小组长' not in r),
            ('二卡剪辑', '🔵', lambda r: '二卡' in r),
            ('剪辑助理', '🟣', lambda r: '助理' in r),
            ('剪辑组长', '🟠', lambda r: '组长' in r and '小组长' not in r),
        ]
        for role_name, icon, matcher in role_labels:
            people_in_role = [n for n in sorted_people if matcher(roles_map[n])]
            if not people_in_role: continue
            tk.Label(cb_frame, text=f'{icon} {role_name}', font=('Microsoft YaHei', 8),
                           bg=C['card'], fg=C['text']).pack(anchor='w', pady=3)
            sub_frame = tk.Frame(cb_frame, bg=C['card'])
            sub_frame.pack(fill='x')
            for nm in people_in_role:
                v = tk.BooleanVar(value=False)
                check_vars[nm] = v
                tk.Checkbutton(sub_frame, text=nm, variable=v, font=('Microsoft YaHei', 8),
                               bg=C['card'], fg=C['text'], selectcolor=C['card'],
                               activebackground=C['card']).pack(side='left', padx=0)

        # 全选/全不选按钮
        sel_frame = tk.Frame(in2, bg=C['card'])
        sel_frame.pack(fill='x', pady=4)
        def _select_all():
            for v in check_vars.values(): v.set(True)
        def _select_none():
            for v in check_vars.values(): v.set(False)
        def _select_card1():
            for n, v in check_vars.items():
                r = roles_map.get(n, '')
                v.set('一卡' in r or '小组长' in r)
        def _select_card2():
            for n, v in check_vars.items():
                r = roles_map.get(n, '')
                v.set('二卡' in r or '助理' in r or ('组长' in r and '小组长' not in r))

        tk.Button(sel_frame, text='全选', font=('Microsoft YaHei', 8),
                  bg=C['blue'], fg='white', relief='flat', cursor='hand2',
                  padx=10, pady=2, command=_select_all).pack(side='left', padx=0)
        tk.Button(sel_frame, text='全不选', font=('Microsoft YaHei', 8),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=10, pady=2, command=_select_none).pack(side='left', padx=0)
        tk.Button(sel_frame, text='小组长+一卡', font=('Microsoft YaHei', 8),
                  bg='#eab308', fg='white', relief='flat', cursor='hand2',
                  padx=10, pady=2, command=_select_card1).pack(side='left', padx=0)
        tk.Button(sel_frame, text='仅二卡/助理/组长', font=('Microsoft YaHei', 8),
                  bg='#d97706', fg='white', relief='flat', cursor='hand2',
                  padx=10, pady=2, command=_select_card2).pack(side='left')

        # 结果展示区
        c3 = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c3.pack(fill='both', expand=True, padx=20, pady=4)
        in3 = tk.Frame(c3, bg=C['card']); in3.pack(fill='both', expand=True, padx=12, pady=6)
        tk.Label(in3, text='📋 分集结果:', font=('Microsoft YaHei', 9, 'bold'),
                 bg=C['card'], fg=C['text']).pack(anchor='w')

        result_text = tk.Text(in3, font=('Consolas', 9), bg=C['log_bg'], fg=C['log_fg'],
                              relief='flat', padx=8, pady=6, height=6, wrap='word')
        result_text.pack(fill='both', expand=True, pady=2)

        # 执行函数
        last_result = [None]  # mutable container

        def _display_result(result, project_name):
            result_text.delete('1.0', 'end')
            result_text.insert('end', f'项目: {project_name}  ({eps_var.get()}集, 一卡区间1～{range_var.get()}集)\n')
            result_text.insert('end', '=' * 60 + '\n\n')
            result_text.insert('end', '📊 统计:\n')
            for k, v in result['stats'].items():
                result_text.insert('end', f'   {k}: {v}\n')
            result_text.insert('end', '\n📋 分集明细 (可直接复制到项目Excel):\n')
            result_text.insert('end', '-' * 60 + '\n')

            selected = [n for n, v in check_vars.items() if v.get()]
            role_labels2 = [
                ('小组长', lambda r: '小组长' in r),
                ('一卡剪辑', lambda r: '一卡' in r and '小组长' not in r),
                ('二卡剪辑', lambda r: '二卡' in r),
                ('剪辑助理', lambda r: '助理' in r),
                ('剪辑组长', lambda r: '组长' in r and '小组长' not in r),
            ]
            for role_name, matcher in role_labels2:
                people = [n for n in selected if matcher(roles_map.get(n, ''))]
                if not people: continue
                result_text.insert('end', f'\n【{role_name}】\n')
                for nm in people:
                    fmt = result['formatted'].get(nm, '')
                    cnt = result['summary'].get(nm, 0)
                    result_text.insert('end', f'  {nm}（{cnt}集）: {fmt}\n')

            result_text.insert('end', '\n\n📝 完整粘贴行 (直接复制到项目Excel分配列):\n')
            result_text.insert('end', '-' * 60 + '\n')
            lines = []
            for nm in selected:
                fmt = result['formatted'].get(nm, '')
                if fmt:
                    lines.append(f'{nm}: {fmt}')
            result_text.insert('end', '\n'.join(lines))

        def _save_txt(proj_name, result):
            """保存分集结果为TXT"""
            try:
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', proj_name)
                txt_path = os.path.join(self.output_dir, f'{safe_name}_分集.txt')
                lines = []
                lines.append(f'项目: {proj_name}  ({eps_var.get()}集, 一卡区间1～{range_var.get()}集)')
                lines.append('=' * 60)
                lines.append('')
                for rn, matcher in [('小组长', lambda r: '小组长' in r),
                                     ('一卡剪辑', lambda r: '一卡' in r and '小组长' not in r),
                                     ('二卡剪辑', lambda r: '二卡' in r),
                                     ('剪辑助理', lambda r: '助理' in r),
                                     ('剪辑组长', lambda r: '组长' in r and '小组长' not in r)]:
                    people = [n for n in result['formatted'] if matcher(roles_map.get(n, ''))]
                    if not people: continue
                    lines.append(f'【{rn}】')
                    for nm in people:
                        lines.append(f'  {nm}（{result["summary"][nm]}集）: {result["formatted"][nm]}')
                    lines.append('')
                lines.append('--- 粘贴行 ---')
                for nm in result['formatted']:
                    f = result['formatted'].get(nm, '')
                    if f: lines.append(f'{nm}: {f}')
                with open(txt_path, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(lines))
                self._log(f'📄 分集TXT已保存: {os.path.basename(txt_path)}')
            except Exception as e:
                self._log(f'⚠️ 保存TXT失败: {e}')

        def _append_to_project(proj_name, result, selected_people, total_eps_val):
            """将分集结果追加到项目数据Excel，格式与现有数据一致"""
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment
                from datetime import datetime, timedelta

                now = datetime.now()
                end_date = now + timedelta(days=1)
                date_str = f'{end_date.month}.{end_date.day}下午18点交'
                # 项目名不带AI前缀，仿照现有格式
                proj_label = proj_name
                dir_label = f'O:\\AI漫剧剪辑一组\\{proj_name}'

                font_title = Font(name='宋体', size=14, bold=True)
                font_normal = Font(name='宋体', size=14)
                align_center = Alignment(horizontal='center', vertical='center')

                wb = load_workbook(self.project_file)
                ws = wb.active

                # 找到末尾，空两行后写入
                last = ws.max_row
                for r in range(last, 0, -1):
                    if ws.cell(r, 1).value or ws.cell(r, 3).value:
                        last = r
                        break
                next_row = last + 2  # 空一行隔开
                title_row = next_row

                # --- 项目标题行 ---
                c = ws.cell(next_row, 1, proj_label)
                c.font = font_title; c.alignment = align_center

                c = ws.cell(next_row, 2, dir_label)
                c.font = font_title; c.alignment = align_center

                c = ws.cell(next_row, 4, date_str)
                c.font = font_title; c.alignment = align_center

                c = ws.cell(next_row, 5, '已分集')
                c.font = font_title; c.alignment = align_center
                next_row += 1

                # --- 人员分配行 ---
                data_start = next_row
                for nm in selected_people:
                    fmt = result['formatted'].get(nm, '')
                    if not fmt: continue
                    c = ws.cell(next_row, 3, f'{nm}：{fmt}')
                    c.font = font_normal; c.alignment = align_center
                    next_row += 1
                data_end = next_row - 1

                # --- 合并单元格 (A/B/D/E列跨所有人员行) ---
                if data_end >= title_row:
                    for col in ['A', 'B', 'D', 'E']:
                        ws.merge_cells(f'{col}{title_row}:{col}{data_end}')

                wb.save(self.project_file)
                wb.close()

                self._log(f'📎 已追加到项目数据: {os.path.basename(self.project_file)}')
                try: os.startfile(self.project_file)
                except: pass
            except Exception as e:
                self._log(f'⚠️ 追加项目数据失败: {e}')

        def _do_assign_common():
            """通用分集逻辑 - do_assign 和 reroll 共用"""
            try:
                _do_assign_inner()
            except Exception as e:
                import traceback
                msg = traceback.format_exc()
                print(msg)
                messagebox.showerror('分集出错', msg)

        def _do_assign_inner():
            """实际分集 + 弹出确认窗"""
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning('提示', '请输入项目名称')
                return
            selected = [n for n, v in check_vars.items() if v.get()]
            if not selected:
                messagebox.showwarning('提示', '请至少选择一位剪辑人员')
                return
            total = eps_var.get()
            rng = range_var.get()

            result = smart_episode_assignment(total, selected, roles_map, rng)
            last_result[0] = result
            _display_result(result, name)

            # 弹出确认窗
            _show_confirm_dialog(dlg, name, selected, total, result,
                                 roles_map, check_vars, last_result,
                                 result_text, eps_var, range_var)

        import re as _re_mod
        import os as _os_mod
        from openpyxl import load_workbook as _load_wb
        from openpyxl.styles import Font as _Font, Alignment as _Alignment
        from datetime import datetime as _dt, timedelta as _td

        # ===== 确认窗构建函数 =====
        def _build_confirm_rows(confirm_dlg, selected, result, outer_frame):
            """在 outer_frame 中为每个 selected 人员创建一行 下拉+输入框"""
            row_widgets = []
            for nm in selected:
                rf = tk.Frame(outer_frame, bg=C['card'])
                rf.pack(fill='x', pady=0)
                nv = tk.StringVar(value=nm)
                cb = ttk.Combobox(rf, textvariable=nv, values=selected,
                                  state='readonly', font=('Microsoft YaHei', 9), width=16)
                cb.pack(side='left', padx=4, pady=3)
                ranges = result['assignments'].get(nm, [])
                parts = [f'{s}-{e}' if s != e else str(s) for s, e in ranges]
                ev = tk.StringVar(value=', '.join(parts))
                ent = tk.Entry(rf, textvariable=ev,
                               font=('Consolas', 9), relief='solid', borderwidth=1)
                ent.pack(side='left', fill='x', expand=True, padx=4, pady=3, ipady=2)
                row_widgets.append((nv, ev, cb, ent))
            return row_widgets

        def _show_confirm_dialog(parent, proj_name, selected, total, result,
                                 roles_map, check_vars, last_result,
                                 result_text, eps_var, range_var):
            cdlg = tk.Toplevel(parent)
            cdlg.title('确认并修改分集结果')
            cdlg.geometry('620x480')
            cdlg.minsize(420, 320)
            cdlg.configure(bg=C['bg'])
            cdlg.transient(parent)
            cdlg.grab_set()

            tk.Label(cdlg, text='📝 确认分集结果（可修改）',
                     font=('Microsoft YaHei', 13, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=(10, 4))
            tk.Label(cdlg, text='下拉选人，编辑集数范围（如 1-3, 5, 7-10），确认后填入表格',
                     font=('Microsoft YaHei', 8), bg=C['bg'], fg=C['text3']).pack()

            outer = tk.Frame(cdlg, bg=C['border'])
            outer.pack(fill='both', expand=True, padx=20, pady=8)

            canvas = tk.Canvas(outer, bg=C['card'], highlightthickness=0)
            sbar = tk.Scrollbar(outer, orient='vertical', command=canvas.yview)
            sframe = tk.Frame(canvas, bg=C['card'])

            sframe.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
            win_id = canvas.create_window((0, 0), window=sframe, anchor='nw')
            canvas.configure(yscrollcommand=sbar.set)

            def _resize(event):
                canvas.itemconfig(win_id, width=event.width)
            canvas.bind('<Configure>', _resize)
            canvas.pack(side='left', fill='both', expand=True)
            sbar.pack(side='right', fill='y')

            # 鼠标滚轮
            def _wheel(event):
                canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
            canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _wheel))
            canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

            # 表头
            hdr = tk.Frame(sframe, bg=C['accent_l'])
            hdr.pack(fill='x')
            tk.Label(hdr, text='剪辑人员', font=('Microsoft YaHei', 8, 'bold'),
                     bg=C['accent_l'], fg=C['text'], width=14, anchor='w', padx=8).pack(side='left')
            tk.Label(hdr, text='负责集数范围', font=('Microsoft YaHei', 8, 'bold'),
                     bg=C['accent_l'], fg=C['text'], anchor='w', padx=8).pack(side='left', fill='x', expand=True)

            row_widgets = _build_confirm_rows(cdlg, selected, result, sframe)

            tk.Label(sframe, text=f'共 {total} 集 | {len(selected)} 人',
                     font=('Microsoft YaHei', 8), bg=C['card'], fg=C['text3'], anchor='e').pack(fill='x', padx=8, pady=(4, 2))

            def _confirm():
                try:
                    lines = []
                    seen = {}
                    for nv, ev, cb, ent in row_widgets:
                        nm = nv.get().strip()
                        eps = ev.get().strip()
                        if not nm or not eps:
                            continue
                        if nm in seen:
                            raise ValueError(f'"{nm}" 出现多次，请合并为一行')
                        seen[nm] = True
                        lines.append(f'{nm}：{eps}')
                    if not lines:
                        messagebox.showwarning('提示', '没有有效的分配行', parent=cdlg)
                        return
                    validated = validate_episode_assignments('\n'.join(lines), selected, total)
                    canvas.unbind_all('<MouseWheel>')
                    cdlg.destroy()
                    last_result[0] = validated
                    _display_result(validated, proj_name)
                    _save_txt(proj_name, validated)
                    _append_to_project(proj_name, validated, selected, total)
                except ValueError as e:
                    messagebox.showerror('校验失败', str(e), parent=cdlg)

            def _on_destroy():
                canvas.unbind_all('<MouseWheel>')
                cdlg.destroy()

            cdlg.protocol('WM_DELETE_WINDOW', _on_destroy)

            btn_f2 = tk.Frame(cdlg, bg=C['bg'])
            btn_f2.pack(fill='x', padx=20, pady=(0, 12))
            tk.Button(btn_f2, text='✅ 确认填入', font=('Microsoft YaHei', 11, 'bold'),
                      bg='#16a34a', fg='white', relief='flat', cursor='hand2',
                      padx=20, pady=6, activebackground='#15803d',
                      command=_confirm).pack(side='left', padx=4)
            tk.Button(btn_f2, text='🔄 再随机一次', font=('Microsoft YaHei', 11),
                      bg=C['amber'], fg='white', relief='flat', cursor='hand2',
                      padx=14, pady=6, activebackground='#b45309',
                      command=lambda: [_on_destroy(),
                                       _do_assign_common()]).pack(side='left', padx=4)
            tk.Button(btn_f2, text='❌ 取消', font=('Microsoft YaHei', 11),
                      bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                      padx=16, pady=6, command=_on_destroy).pack(side='right', padx=4)

        def copy_result():
            text = result_text.get('1.0', 'end-1c')
            if text.strip():
                dlg.clipboard_clear()
                dlg.clipboard_append(text)
                messagebox.showinfo('已复制', '分集结果已复制到剪贴板，可直接粘贴到项目Excel。')

        btn_f = tk.Frame(dlg, bg=C['bg'])
        btn_f.pack(fill='x', padx=20, pady=0)
        tk.Button(btn_f, text='🎲 随机分集', font=('Microsoft YaHei', 12, 'bold'),
                  bg='#e11d48', fg='white', relief='flat', cursor='hand2',
                  padx=24, pady=8, activebackground='#be123c',
                  command=_do_assign_common).pack(side='left', padx=0)
        tk.Button(btn_f, text='🎲 再次随机', font=('Microsoft YaHei', 11),
                  bg='#d97706', fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=8, activebackground='#b45309',
                  command=_do_assign_common).pack(side='left', padx=0)
        tk.Button(btn_f, text='📋 复制结果', font=('Microsoft YaHei', 11),
                  bg=C['blue'], fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=8, command=copy_result).pack(side='left', padx=0)
        tk.Button(btn_f, text='关闭', font=('Microsoft YaHei', 11),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=8, command=dlg.destroy).pack(side='right')

    def _check_duplicates(self):
        self._log('🔍 正在检查项目去重...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, _ = gc.parse_projects(df)
            if not records:
                self._log('⚠️ 未解析到有效记录')
                _sys.path.pop(0)
                return

            ok = gc.self_check(records)
            id_names = {}
            for r in records:
                pid = r['项目ID']
                if pid:
                    if pid not in id_names:
                        id_names[pid] = set()
                    id_names[pid].add(r['AI项目名称'][:30])

            issues = {pid: names for pid, names in id_names.items() if len(names) > 1}
            if issues:
                msg = f'⚠️ 发现 {len(issues)} 个项目ID存在名称不一致:\n'
                for pid, names in list(issues.items())[:10]:
                    msg += f'  ID={pid}: {names}\n'
                self._log(msg)
                messagebox.showwarning('去重检查', msg)
            else:
                self._log(f'✅ 全部 {len(id_names)} 个项目ID名称一致，无重复。')
                messagebox.showinfo('去重检查', f'✅ 检查通过！\n\n{len(id_names)} 个项目ID，名称均一致，无重复。')
            _sys.path.pop(0)
        except Exception as e:
            self._log(f'❌ 检查失败: {e}')
            messagebox.showerror('失败', str(e))

    # ============ 新增功能：数据校验 ============
    def _validate_data(self):
        self._log('🔍 正在校验项目数据...')
        try:
            issues = validate_project_data(self.project_file, self.cfg.get('人员角色', {}))
            if not issues:
                self._log('✅ 数据校验通过，无问题！')
                messagebox.showinfo('校验通过', '✅ 项目数据校验通过！\n\n未发现日期异常、人名不匹配、或集数重复分配等问题。')
                return

            self._log(f'⚠️ 发现 {len(issues)} 个问题:')
            for loc, item, detail in issues[:20]:
                self._log(f'  [{loc}] {item}: {detail}')

            dlg = tk.Toplevel(self.root)
            dlg.title('数据校验结果')
            dlg.geometry('650x400')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text=f'⚠️ 发现 {len(issues)} 个问题', font=('Microsoft YaHei', 14, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=12)
            txt = tk.Text(dlg, font=('Consolas', 10), bg=C['log_bg'], fg=C['log_fg'],
                          relief='flat', padx=10, pady=8)
            for loc, item, detail in issues[:50]:
                txt.insert('end', f'[{loc}] {item}: {detail}\n')
            txt.configure(state='disabled')
            txt.pack(fill='both', expand=True, padx=20, pady=0)
        except Exception as e:
            self._log(f'❌ 校验失败: {e}')

    # ============ 新增功能：个人趋势 ============
    def _gen_trend(self):
        name = self._trend_var.get()
        if not name:
            messagebox.showwarning('提示', '请选择人员'); return
        self._log(f'📊 正在生成 {name} 的月度趋势...')
        try:
            roles_map = self.cfg.get('人员角色', {})
            path = generate_person_trend_html(name, roles_map, self.output_dir)
            if path:
                self._log(f'✅ 趋势图已生成: {os.path.basename(path)}')
                try: os.startfile(path)
                except: pass
            else:
                self._log('⚠️ 未找到该人员的历史数据（可能需要先生成几个月的提成表）')
                messagebox.showinfo('提示', '未找到历史数据。\n\n请确保在输出目录中有多个以 "AI后期剪辑提成一组" 开头的 Excel 文件。')
        except Exception as e:
            self._log(f'❌ 趋势生成失败: {e}')

    # ============ 功能4：高级查询与筛选 ============
    def _advanced_filter(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('🔎 高级查询与筛选')
        dlg.geometry('850x650')
        dlg.minsize(700, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='🔎 高级查询与筛选', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=14)
        tk.Label(dlg, text='按条件组合筛选，支持导出筛选结果', font=('Microsoft YaHei', 9),
                 bg=C['bg'], fg=C['text3']).pack(pady=0)

        # 筛选条件区
        cf = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        cf.pack(fill='x', padx=16, pady=0)
        cfi = tk.Frame(cf, bg=C['card']); cfi.pack(fill='x', padx=12, pady=8)

        # 第一行：姓名 + 角色
        r1 = tk.Frame(cfi, bg=C['card']); r1.pack(fill='x', pady=2)
        tk.Label(r1, text='姓名:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        name_entry = tk.Entry(r1, font=('Microsoft YaHei', 10), width=14)
        name_entry.pack(side='left', padx=0)

        tk.Label(r1, text='角色:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        role_vars = {}
        for role in ROLES:
            v = tk.BooleanVar(value=True)
            role_vars[role] = v
            tk.Checkbutton(r1, text=role, variable=v, font=('Microsoft YaHei', 8),
                           bg=C['card'], fg=C['text'], selectcolor=C['card'],
                           activebackground=C['card']).pack(side='left', padx=0)

        # 第二行：集数范围 + 项目ID
        r2 = tk.Frame(cfi, bg=C['card']); r2.pack(fill='x', pady=2)
        tk.Label(r2, text='集数范围:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        eps_min = tk.Entry(r2, font=('Microsoft YaHei', 10), width=6)
        eps_min.pack(side='left')
        tk.Label(r2, text='~', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=2)
        eps_max = tk.Entry(r2, font=('Microsoft YaHei', 10), width=6)
        eps_max.pack(side='left', padx=0)

        tk.Label(r2, text='项目ID:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        pid_entry = tk.Entry(r2, font=('Microsoft YaHei', 10), width=10)
        pid_entry.pack(side='left', padx=0)

        tk.Label(r2, text='绩效:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=6, anchor='w').pack(side='left')
        status_var = tk.StringVar(value='全部')
        ttk.Combobox(r2, textvariable=status_var, values=['全部', '是', '否'],
                     state='readonly', font=('Microsoft YaHei', 10), width=6).pack(side='left', padx=0)

        tk.Label(r2, text='提成范围:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        comm_min = tk.Entry(r2, font=('Microsoft YaHei', 10), width=7)
        comm_min.pack(side='left')
        tk.Label(r2, text='~', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=2)
        comm_max = tk.Entry(r2, font=('Microsoft YaHei', 10), width=7)
        comm_max.pack(side='left')

        # 按钮
        btn_f = tk.Frame(cfi, bg=C['card']); btn_f.pack(fill='x', pady=8)

        # 结果区
        result_frame = tk.Frame(dlg, bg=C['bg'])
        result_frame.pack(fill='both', expand=True, padx=16, pady=0)

        cols = ('姓名', '角色', '集数', '项目数', '基准', '绩效', '提成')
        tree = ttk.Treeview(result_frame, columns=cols, show='headings', height=14)
        col_widths = [90, 95, 65, 60, 60, 55, 85]
        for c, w in zip(cols, col_widths):
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor='center')
        tree.pack(side='left', fill='both', expand=True)
        tree.tag_configure('ok', foreground='#16a34a')
        tree.tag_configure('fail', foreground='#dc2626')

        sb = ttk.Scrollbar(result_frame, orient='vertical', command=tree.yview)
        sb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=sb.set)

        def _int_or_none(s):
            try: return int(s.strip())
            except: return None

        def _do_filter():
            for item in tree.get_children():
                tree.delete(item)
            try:
                import pandas as pd
                gc, _sys = self._load_gc_module()
                df = pd.read_excel(self.project_file, header=None)
                records, group_pids = gc.parse_projects(df)
                cd = gc.compute_commission(records, group_pids)
                _sys.path.pop(0)

                filters = {
                    'name_keyword': name_entry.get().strip() or None,
                    'roles': [r for r, v in role_vars.items() if v.get()],
                    'min_eps': _int_or_none(eps_min.get()),
                    'max_eps': _int_or_none(eps_max.get()),
                    'project_id': pid_entry.get().strip() or None,
                    'status': status_var.get(),
                    'min_commission': _int_or_none(comm_min.get()),
                    'max_commission': _int_or_none(comm_max.get()),
                }
                results = advanced_filter(records, cd, filters)

                if not results:
                    self._log('🔎 筛选结果：0 条匹配')
                    return

                total_eps = sum(r['episodes'] for r in results)
                total_comm = sum(r['commission'] for r in results)
                self._log(f'🔎 筛选结果：{len(results)} 人，总集数 {total_eps}，总提成 {total_comm:,}')

                for r in results:
                    tag = 'ok' if r['status'] == '是' else 'fail'
                    tree.insert('', 'end', values=(
                        r['name'], r['role'], r['episodes'], r['projects'],
                        f'{r["quota"]}集' if r['quota'] > 0 else '无', r['status'],
                        f'{r["commission"]:,}'), tags=(tag,))
            except Exception as e:
                self._log(f'❌ 筛选失败: {e}')

        def _export_csv():
            items = tree.get_children()
            if not items:
                messagebox.showwarning('提示', '无数据可导出，请先执行筛选')
                return
            import csv
            path = filedialog.asksaveasfilename(
                title='导出筛选结果',
                defaultextension='.csv',
                filetypes=[('CSV文件', '*.csv')],
                initialdir=self.output_dir,
                initialfile='筛选结果.csv')
            if not path: return
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(cols)
                for item in items:
                    writer.writerow(tree.item(item)['values'])
            self._log(f'📥 已导出: {os.path.basename(path)}')
            try: os.startfile(os.path.dirname(path))
            except: pass

        tk.Button(btn_f, text='🔍 执行筛选', font=('Microsoft YaHei', 11, 'bold'),
                  bg=C['accent'], fg='white', relief='flat', cursor='hand2',
                  padx=20, pady=6, command=_do_filter).pack(side='left', padx=0)
        tk.Button(btn_f, text='📥 导出CSV', font=('Microsoft YaHei', 10),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=6, command=_export_csv).pack(side='left', padx=0)
        tk.Button(btn_f, text='关闭', font=('Microsoft YaHei', 10),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=6, command=dlg.destroy).pack(side='right')

        name_entry.bind('<Return>', lambda e: _do_filter())

    # ============ 功能5：数据修正工具 ============
    def _data_correction(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('✏️ 数据修正工具')
        dlg.geometry('900x650')
        dlg.minsize(750, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='✏️ 数据修正工具', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=14)
        tk.Label(dlg, text='双击某行可修正集数或删除。修改后自动重算提成并更新Excel。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text3']).pack(pady=0)

        # 数据加载
        cols = ('#', '姓名', '角色', '项目ID', '项目名', '集数', '明细')
        tree = ttk.Treeview(dlg, columns=cols, show='headings', height=16)
        col_widths = [35, 80, 85, 70, 200, 55, 140]
        for c, w in zip(cols, col_widths):
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor='center')
        tree.pack(fill='both', expand=True, padx=16, pady=0)

        self._correction_records = []
        self._correction_gc = None
        self._correction_sys = None

        def _load_data():
            for item in tree.get_children():
                tree.delete(item)
            try:
                import pandas as pd
                gc, _sys = self._load_gc_module()
                df = pd.read_excel(self.project_file, header=None)
                records, group_pids = gc.parse_projects(df)
                self._correction_records = records
                self._correction_gc = gc
                self._correction_sys = _sys

                for i, r in enumerate(records):
                    pname = r['AI项目名称'][:35] if r['AI项目名称'] else ''
                    detail = r['完成明细'][:25] if r['完成明细'] else ''
                    tree.insert('', 'end', iid=str(i), values=(
                        i + 1, r['身份证姓名'], r['角色'],
                        r['项目ID'], pname, r['单项目数/集数'], detail))
                self._log(f'✏️ 已加载 {len(records)} 条记录，双击可编辑')
            except Exception as e:
                self._log(f'❌ 加载失败: {e}')

        _load_data()

        def _edit_record(event):
            sel = tree.selection()
            if not sel: return
            idx = int(sel[0])
            rec = self._correction_records[idx]

            ed = tk.Toplevel(dlg)
            ed.title('修正记录')
            ed.geometry('420x310')
            ed.configure(bg=C['bg'])
            ed.transient(dlg); ed.grab_set()

            tk.Label(ed, text=f'修正记录 #{idx+1}', font=('Microsoft YaHei', 13, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=14)

            cf = tk.Frame(ed, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
            cf.pack(fill='x', padx=20, pady=0)
            ci = tk.Frame(cf, bg=C['card']); ci.pack(fill='x', padx=12, pady=10)

            fields = [
                ('姓名:', 'name', rec['身份证姓名']),
                ('项目ID:', 'pid', rec['项目ID']),
                ('项目名:', 'pname', rec['AI项目名称']),
                ('集数:', 'eps', str(rec['单项目数/集数'])),
            ]
            entries = {}
            for j, (label, key, val) in enumerate(fields):
                tk.Label(ci, text=label, font=('Microsoft YaHei', 9), bg=C['card'],
                         fg=C['text'], width=8, anchor='w').grid(row=j, column=0, sticky='w', pady=3)
                e = tk.Entry(ci, font=('Microsoft YaHei', 10), width=35)
                e.insert(0, val)
                e.grid(row=j, column=1, padx=6, pady=3)
                entries[key] = e

            def _save_edit():
                rec['身份证姓名'] = entries['name'].get().strip()
                rec['项目ID'] = entries['pid'].get().strip()
                rec['AI项目名称'] = entries['pname'].get().strip()
                try:
                    new_eps = int(entries['eps'].get())
                    rec['单项目数/集数'] = new_eps
                    rec['完成明细'] = ','.join(str(i) for i in range(1, new_eps + 1))
                except:
                    pass
                tree.set(str(idx), column='姓名', value=rec['身份证姓名'])
                tree.set(str(idx), column='项目ID', value=rec['项目ID'])
                tree.set(str(idx), column='集数', value=rec['单项目数/集数'])
                tree.set(str(idx), column='明细', value=rec['完成明细'][:25])
                self._log(f'✏️ 已修正记录 #{idx+1}')
                ed.destroy()

            def _delete_rec():
                if messagebox.askyesno('确认删除', f'确定要删除记录 #{idx+1} 吗？\n\n此操作不可恢复。'):
                    self._correction_records.pop(idx)
                    tree.delete(str(idx))
                    self._log(f'🗑️ 已删除记录 #{idx+1}')
                    ed.destroy()
                    _load_data()

            bf = tk.Frame(ed, bg=C['bg']); bf.pack(fill='x', padx=20, pady=0)
            tk.Button(bf, text='💾 保存修改', font=('Microsoft YaHei', 11, 'bold'),
                      bg=C['green'], fg='white', relief='flat', cursor='hand2',
                      padx=20, pady=6, command=_save_edit).pack(side='left', padx=0)
            tk.Button(bf, text='🗑️ 删除', font=('Microsoft YaHei', 11),
                      bg=C['red'], fg='white', relief='flat', cursor='hand2',
                      padx=16, pady=6, command=_delete_rec).pack(side='left')
            tk.Button(bf, text='取消', font=('Microsoft YaHei', 11),
                      bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                      padx=16, pady=6, command=ed.destroy).pack(side='right')

        tree.bind('<Double-1>', _edit_record)

        def _regenerate():
            if not self._correction_records:
                messagebox.showwarning('提示', '无数据可生成'); return
            if not messagebox.askyesno('确认重新生成', '将使用修正后的数据重新生成提成表，确定吗？'):
                return
            try:
                gc = self._correction_gc
                # 重新计算
                group_pids = {g: set() for g in gc.GROUPS}
                for r in self._correction_records:
                    pid = r['项目ID']
                    for gname, ginfo in gc.GROUPS.items():
                        if r['身份证姓名'] in ginfo['成员'] and pid:
                            group_pids[gname].add(pid)

                cd = gc.compute_commission(self._correction_records, group_pids)
                excel_path = os.path.join(self.output_dir,
                    f'AI后期剪辑提成一组{self._correction_gc.OUTPUT_MONTH}.xlsx')
                path, html_path = gc.generate_excel(
                    self._correction_records, cd,
                    self._correction_gc.TEMPLATE_FILE, excel_path)
                self._log(f'✅ 已用修正数据重新生成')
                if self._correction_sys:
                    self._correction_sys.path.pop(0)
                dlg.destroy()
                try: os.startfile(path)
                except: pass
                try: os.startfile(html_path)
                except: pass
            except Exception as e:
                self._log(f'❌ 重新生成失败: {e}')

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=16, pady=0)
        tk.Button(bf, text='🔄 重新加载', font=('Microsoft YaHei', 10),
                  bg=C['blue'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=5, command=_load_data).pack(side='left', padx=0)
        tk.Button(bf, text='▶ 用修正数据重新生成', font=('Microsoft YaHei', 11, 'bold'),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=5, command=_regenerate).pack(side='left', padx=0)
        tk.Button(bf, text='关闭', font=('Microsoft YaHei', 10),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=5, command=dlg.destroy).pack(side='right')

    # ============ 功能7：项目数据模板下载 ============
    def _download_template(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return
        try:
            path = filedialog.asksaveasfilename(
                title='保存模板到',
                defaultextension='.xlsx',
                filetypes=[('Excel文件', '*.xlsx')],
                initialdir=self.output_dir,
                initialfile='AI项目数据录入模板.xlsx')
            if not path: return
            generate_project_template(path)
            self._log(f'📥 模板已生成: {os.path.basename(path)}')
            try: os.startfile(path)
            except: pass
            messagebox.showinfo('完成', f'✅ 项目数据录入模板已生成！\n\n📄 {os.path.basename(path)}\n\n包含：\n• 示例数据行\n• 录入规范说明Sheet')
        except Exception as e:
            self._log(f'❌ 模板生成失败: {e}')
            messagebox.showerror('失败', str(e))

    # ============ 功能8：配置快照与回滚 ============
    def _config_snapshot(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return
        snap_dir = os.path.join(SCRIPT_DIR, 'config_snapshots')
        snapshots = list_config_snapshots(snap_dir)

        dlg = tk.Toplevel(self.root)
        dlg.title('📸 配置快照管理')
        dlg.geometry('650x500')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='📸 配置快照管理', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=14)
        tk.Label(dlg, text='保存配置历史版本，支持一键回滚', font=('Microsoft YaHei', 9),
                 bg=C['bg'], fg=C['text3']).pack(pady=0)

        # 列表
        list_frame = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=0)

        cols = ('快照文件', '时间')
        tree = ttk.Treeview(list_frame, columns=cols, show='headings', height=12)
        tree.heading('快照文件', text='快照文件')
        tree.heading('时间', text='时间')
        tree.column('快照文件', width=350, anchor='w')
        tree.column('时间', width=180, anchor='center')
        tree.pack(side='left', fill='both', expand=True, padx=4, pady=4)

        sb = ttk.Scrollbar(list_frame, orient='vertical', command=tree.yview)
        sb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=sb.set)

        def _refresh_list():
            for item in tree.get_children():
                tree.delete(item)
            snaps = list_config_snapshots(snap_dir)
            for s in snaps:
                tree.insert('', 'end', values=(s['name'], s['mtime'].strftime('%Y-%m-%d %H:%M:%S')))

        _refresh_list()

        def _create_snap():
            path = create_config_snapshot(CONFIG_PATH, snap_dir)
            self._log(f'📸 配置快照已保存: {os.path.basename(path)}')
            _refresh_list()
            messagebox.showinfo('完成', f'✅ 快照已保存！\n\n{os.path.basename(path)}')

        def _restore_snap():
            sel = tree.selection()
            if not sel: return
            val = tree.item(sel[0])['values']
            fname = val[0]
            if not messagebox.askyesno('确认回滚', f'确定要恢复到快照 "{fname}" 吗？\n\n当前配置将被覆盖。'):
                return
            restore_config_snapshot(os.path.join(snap_dir, fname), CONFIG_PATH)
            self.cfg = self._load_config()
            self._refresh_role_tags()
            self._log(f'📸 已回滚到快照: {fname}')
            messagebox.showinfo('完成', f'✅ 已回滚到快照！\n\n{fname}\n\n配置已更新，下次生成时生效。')
            _refresh_list()

        def _delete_snap():
            sel = tree.selection()
            if not sel: return
            val = tree.item(sel[0])['values']
            if messagebox.askyesno('确认删除', f'确定要删除快照 "{val[0]}" 吗？'):
                try:
                    os.remove(os.path.join(snap_dir, val[0]))
                    _refresh_list()
                except Exception as e:
                    messagebox.showerror('失败', str(e))

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=20, pady=0)
        tk.Button(bf, text='📸 创建快照', font=('Microsoft YaHei', 11, 'bold'),
                  bg=C['accent'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=6, command=_create_snap).pack(side='left', padx=0)
        tk.Button(bf, text='↩ 回滚到此', font=('Microsoft YaHei', 11),
                  bg=C['orange'], fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=6, command=_restore_snap).pack(side='left', padx=0)
        tk.Button(bf, text='🗑️ 删除', font=('Microsoft YaHei', 11),
                  bg=C['red'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=6, command=_delete_snap).pack(side='left', padx=0)
        tk.Button(bf, text='关闭', font=('Microsoft YaHei', 11),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=6, command=dlg.destroy).pack(side='right')

    # ============ 功能9：Web仪表盘服务 ============
    def _web_server(self):
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return

        if hasattr(self, '_web_server_instance') and self._web_server_instance:
            self._web_server_instance.shutdown()
            self._web_server_instance.server_close()
            self._web_server_instance = None
            self._log('🛑 Web服务已停止')
            return

        try:
            self._web_server_instance = start_web_server(self.output_dir, 8080)
            port = self._web_server_instance.server_address[1]
            self._log(f'🌐 Web仪表盘已启动: http://localhost:{port}')
            self._log(f'   服务目录: {self.output_dir}')
            self._log(f'   浏览器访问查看全部报表和卡片')

            import webbrowser
            webbrowser.open(f'http://localhost:{port}')

            # 后台运行
            def serve():
                self._web_server_instance.serve_forever()

            threading.Thread(target=serve, daemon=True).start()
            messagebox.showinfo('Web服务已启动',
                f'✅ Web仪表盘已启动！\n\n'
                f'📍 地址: http://localhost:{port}\n'
                f'📂 目录: {os.path.basename(self.output_dir)}\n\n'
                f'浏览器已自动打开。\n'
                f'请勿关闭此窗口。')
        except Exception as e:
            self._log(f'❌ Web服务启动失败: {e}')
            messagebox.showerror('失败', str(e))

    # 功能10：移动端适配绩效卡片 —— 在 features.py 中改进 generate_person_cards
    def _refresh_backups(self):
        if not HAS_FEATURES: return
        backups = list_backups(BACKUP_DIR)
        self._backup_list.configure(state='normal')
        self._backup_list.delete('1.0', 'end')
        if not backups:
            self._backup_list.insert('1.0', '暂无备份文件')
        else:
            total_size = sum(b['size'] for b in backups)
            self._backup_list.insert('1.0', f'共 {len(backups)} 个备份文件，占用 {total_size/1024:.1f} KB\n')
            self._backup_list.insert('end', '-' * 50 + '\n')
            for b in backups[:30]:
                self._backup_list.insert('end',
                    f'{b["mtime"].strftime("%m/%d %H:%M")}  {b["size"]/1024:6.1f}KB  {b["name"]}\n')
        self._backup_list.configure(state='disabled')

    def _manage_backups(self):
        if not HAS_FEATURES: return
        removed = cleanup_backups(BACKUP_DIR, keep=30)
        self._log(f'🗄️ 已清理 {removed} 个旧备份，保留最近30个')
        self._refresh_backups()
        messagebox.showinfo('完成', f'已清理 {removed} 个旧备份文件。\n当前保留最近30个。')

    def check_files(self):
        items = [
            (self.project_file, '项目数据'),
            (CONFIG_PATH, '角色配置'),
            (self.template_file, '模板文件'),
        ]
        ok = True
        for fpath, desc in items:
            if not os.path.exists(fpath):
                ok = False
        self.run_btn.configure(state='normal' if ok else 'disabled',
                                bg=C['green'] if ok else C['gray'])
        if ok:
            self._log('✅ 文件就绪，点击"一键生成"开始。')
        else:
            self._log('❌ 文件不完整，请检查。')

    # ============ 日志 ============

    def _log(self, msg):
        self.log_txt.configure(state='normal')
        self.log_txt.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.log_txt.see('end'); self.log_txt.configure(state='disabled')

    def _filter_line(self, line):
        line = line.strip()
        if not line or line.startswith('PS '):
            return
        skip = ['所在位置', 'CategoryInfo', 'FullyQualifiedErrorId', 'NativeCommandError',
                'RemoteException', '~~~~~~~~~~', '鎸変换鎰忛', 'EOFError']
        if any(s in line for s in skip): return
        self._log(line)

    # ============ 生成 ============

    def run(self):
        """第一步：解析项目 → 第二步：弹出超时对话框 → 第三步：生成"""
        self.run_btn.configure(state='disabled', bg=C['gray'], text='⏳ 解析中...')
        self.st.configure(text='⏳ 正在解析项目数据...')
        self.progress.start(10)
        self._log('🔍 第一步：解析项目数据...')
        self._log('—' * 50)

        project_file = self.project_file
        output_dir = self.output_dir
        template_file = self.template_file

        def parse_worker():
            try:
                import pandas as pd
                gc, _sys = self._load_gc_module()

                df = pd.read_excel(project_file, header=None)
                self._log(f'   项目文件: {os.path.basename(project_file)} ({len(df)}行)')

                # 先用不含超时的解析收集项目信息
                records, group_pids = gc.parse_projects(df)
                _sys.path.pop(0)

                # 收集每个项目的集数范围
                proj_info = {}  # pid -> {name, eps_set, people}
                for r in records:
                    pid = r['项目ID']
                    if not pid:
                        continue
                    if pid not in proj_info:
                        proj_info[pid] = {
                            'name': r['AI项目名称'][:40],
                            'eps_set': set(),
                            'people': {}
                        }
                    detail = r['完成明细']
                    for part in detail.split(','):
                        part = part.strip()
                        if part.isdigit():
                            proj_info[pid]['eps_set'].add(int(part))
                    nm = r['身份证姓名']
                    if nm not in proj_info[pid]['people']:
                        proj_info[pid]['people'][nm] = set()
                    for part in detail.split(','):
                        part = part.strip()
                        if part.isdigit():
                            proj_info[pid]['people'][nm].add(int(part))

                self.root.after(0, lambda: self._show_overtime_dialog(
                    proj_info, project_file, output_dir, template_file))
            except Exception as e:
                self.root.after(0, lambda: self._log(f'❌ 解析失败: {e}'))
                import traceback
                self.root.after(0, lambda: self._log(traceback.format_exc()))
                self.root.after(0, self._done)

        threading.Thread(target=parse_worker, daemon=True).start()

    def _show_overtime_dialog(self, proj_info, project_file, output_dir, template_file):
        """弹出超时集数标记对话框：搜索 → 选中 → 输入集数 → 确认"""
        self.progress.stop()
        self.run_btn.configure(state='normal', bg=C['accent'], text='▶  一键生成提成表')
        self._log(f'📋 已解析 {len(proj_info)} 个项目，请标记超4分钟的集数')

        dlg = tk.Toplevel(self.root)
        dlg.title('⏱️ 标记超4分钟集数')
        dlg.geometry('860x620')
        dlg.minsize(720, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root)
        dlg.grab_set()

        # ---- 标题 ----
        hdr_f = tk.Frame(dlg, bg=C['hdr_bg'])
        hdr_f.pack(fill='x')
        tk.Label(hdr_f, text='⏱️  标记超4分钟集数（超时算2集）',
                 font=('Microsoft YaHei', 14, 'bold'), fg=C['hdr_text'],
                 bg=C['hdr_bg']).pack(pady=10)

        # ---- 搜索栏 + 已标记摘要 ----
        top_bar = tk.Frame(dlg, bg=C['bg'])
        top_bar.pack(fill='x', padx=12, pady=(8, 4))

        tk.Label(top_bar, text='🔍 搜索:', font=('Microsoft YaHei', 10),
                 bg=C['bg'], fg=C['text']).pack(side='left', padx=(0, 6))
        search_var = tk.StringVar()
        search_entry = tk.Entry(top_bar, textvariable=search_var,
                                font=('Microsoft YaHei', 11), width=30,
                                relief='solid', borderwidth=1)
        search_entry.pack(side='left', padx=(0, 8))

        summary_frame = tk.Frame(top_bar, bg=C['bg'])
        summary_frame.pack(side='left', fill='x', expand=True)
        summary_label = tk.Label(summary_frame, text='',
                                 font=('Microsoft YaHei', 8), bg=C['bg'],
                                 fg=C['text2'], anchor='w')
        summary_label.pack(side='left')

        # ---- 主内容区：左列表 + 右输入 ----
        main_frame = tk.Frame(dlg, bg=C['bg'])
        main_frame.pack(fill='both', expand=True, padx=12, pady=(4, 0))

        # 左：项目列表
        left_panel = tk.Frame(main_frame, bg=C['card'], highlightthickness=1,
                              highlightbackground=C['border'], width=300)
        left_panel.pack(side='left', fill='both', padx=(0, 6))
        left_panel.pack_propagate(False)

        tk.Label(left_panel, text='项目列表', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(padx=10, pady=(8, 4), anchor='w')

        list_frame = tk.Frame(left_panel, bg=C['card'])
        list_frame.pack(fill='both', expand=True, padx=6, pady=(0, 6))

        listbox = tk.Listbox(list_frame, font=('Microsoft YaHei', 9),
                             bg='white', fg=C['text'], relief='flat',
                             highlightthickness=1, highlightbackground=C['border'],
                             selectbackground=C['accent'], selectforeground='white',
                             activestyle='none')
        list_scroll = ttk.Scrollbar(list_frame, orient='vertical', command=listbox.yview)
        listbox.configure(yscrollcommand=list_scroll.set)
        listbox.pack(side='left', fill='both', expand=True)
        list_scroll.pack(side='right', fill='y')

        # 右：输入面板
        right_panel = tk.Frame(main_frame, bg=C['card'], highlightthickness=1,
                               highlightbackground=C['border'])
        right_panel.pack(side='left', fill='both', expand=True)

        tk.Label(right_panel, text='超时集数输入', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(padx=12, pady=(8, 4), anchor='w')

        selected_info = tk.Label(right_panel, text='← 请先在左侧搜索并点击项目',
                                 font=('Microsoft YaHei', 9), bg=C['card'],
                                 fg=C['text3'], justify='left')
        selected_info.pack(padx=12, pady=(0, 6), anchor='w')

        input_frame = tk.Frame(right_panel, bg=C['card'])
        input_frame.pack(fill='x', padx=12, pady=(0, 4))

        tk.Label(input_frame, text='超4分钟的集数:',
                 font=('Microsoft YaHei', 9, 'bold'), bg=C['card'],
                 fg=C['text']).pack(anchor='w', pady=(0, 4))

        ot_var = tk.StringVar()
        ot_entry = tk.Entry(input_frame, textvariable=ot_var,
                            font=('Consolas', 11), width=35,
                            relief='solid', borderwidth=1)
        ot_entry.pack(fill='x', pady=(0, 4))

        tk.Label(input_frame, text='支持: 1,3,5-8,10  或  1 3 5-8 10（空格/逗号/分号均可）',
                 font=('Microsoft YaHei', 7), bg=C['card'],
                 fg=C['text3']).pack(anchor='w')

        preview_label = tk.Label(input_frame, text='',
                                 font=('Microsoft YaHei', 8), bg=C['card'],
                                 fg=C['orange'], anchor='w')
        preview_label.pack(fill='x', pady=(4, 0))

        save_btn = tk.Button(right_panel, text='💾 保存此项目标记',
                             font=('Microsoft YaHei', 10, 'bold'),
                             bg=C['green'], fg='white', relief='flat',
                             cursor='hand2', padx=16, pady=6,
                             activebackground=C['green_l'])
        save_btn.pack(padx=12, pady=(8, 12), anchor='w')

        # ---- 数据模型 ----
        overtime_data = {}   # {pid: set of int}
        current_pid = [None]

        all_projects = []
        for pid in sorted(proj_info.keys(), key=lambda p: int(p) if p.isdigit() else 0):
            info = proj_info[pid]
            all_projects.append((f'{pid}  {info["name"]}', pid, info))

        def _populate_list(filter_text=''):
            listbox.delete(0, 'end')
            ft = filter_text.strip().lower()
            for disp, pid, info in all_projects:
                if ft and ft not in disp.lower():
                    continue
                marker = ' ●' if pid in overtime_data and overtime_data[pid] else ''
                listbox.insert('end', disp + marker)
                if pid in overtime_data and overtime_data[pid]:
                    listbox.itemconfig('end', bg='#fef3c7')

        def _update_summary():
            parts = []
            for pid in sorted(overtime_data.keys(), key=lambda p: int(p) if p.isdigit() else 0):
                eps = sorted(overtime_data[pid])
                parts.append(f'{pid}({",".join(str(e) for e in eps)})')
            summary_label.configure(
                text='已标记: ' + ' | '.join(parts) if parts else '暂未标记任何超时集数')

        def _on_list_select(event):
            sel = listbox.curselection()
            if not sel: return
            target_idx = sel[0]
            ft = search_var.get().strip().lower()
            count = 0
            for disp, pid, info in all_projects:
                if ft and ft not in disp.lower():
                    continue
                if count == target_idx:
                    current_pid[0] = pid
                    selected_info.configure(
                        text=f'📁 ID: {pid}\n📝 {info["name"]}\n📊 总集数: {len(info["eps_set"])}集',
                        fg=C['text'])
                    if pid in overtime_data:
                        eps = sorted(overtime_data[pid])
                        ot_var.set(','.join(str(e) for e in eps))
                        _show_preview(eps)
                    else:
                        ot_var.set('')
                        preview_label.configure(text='')
                    return
                count += 1

        def _show_preview(eps_list=None):
            if eps_list is not None:
                preview_label.configure(text=f'当前: {len(eps_list)}集 → 实际算 {len(eps_list)*2}集', fg=C['orange'])
                return
            text = ot_var.get().strip()
            if not text:
                preview_label.configure(text='')
                return
            try:
                from features import parse_overtime_episodes
                eps = parse_overtime_episodes(text)
                if eps:
                    preview_label.configure(text=f'解析: {len(eps)}集 → {sorted(eps)[:15]} 实际算{len(eps)*2}集', fg=C['orange'])
                else:
                    preview_label.configure(text='未识别到有效集数', fg=C['red'])
            except:
                preview_label.configure(text='格式错误', fg=C['red'])

        def _on_input_change(*args):
            _show_preview()

        def _save_current():
            pid = current_pid[0]
            if not pid:
                messagebox.showwarning('提示', '请先在左侧列表中选择一个项目')
                return
            text = ot_var.get().strip()
            if not text:
                overtime_data.pop(pid, None)
            else:
                try:
                    from features import parse_overtime_episodes
                    eps = parse_overtime_episodes(text)
                    if not eps:
                        messagebox.showwarning('提示', '未能识别到有效的集数，请检查格式')
                        return
                    overtime_data[pid] = eps
                except Exception as e:
                    messagebox.showwarning('错误', f'解析失败: {e}')
                    return
            _update_summary()
            _populate_list(search_var.get())
            self._log(f'⏱️ 项目 {pid}: {text if text else "(已清除)"}')

        def _on_search_change(*args):
            _populate_list(search_var.get())

        search_var.trace_add('write', _on_search_change)
        ot_var.trace_add('write', _on_input_change)
        listbox.bind('<<ListboxSelect>>', _on_list_select)
        save_btn.configure(command=_save_current)
        search_entry.bind('<Return>', lambda e: listbox.focus_set())

        _populate_list()

        # ---- 底部按钮 ----
        btn_frame = tk.Frame(dlg, bg=C['bg'])
        btn_frame.pack(fill='x', padx=12, pady=12)

        def _confirm():
            save_data = {pid: sorted(eps) for pid, eps in overtime_data.items() if eps}
            total_ot = sum(len(v) for v in save_data.values())
            self._log(f'⏱️ 已标记 {total_ot} 个超时集数 ({len(save_data)}个项目)')
            dlg.destroy()
            self._do_generate(project_file, output_dir, template_file, save_data)

        def _skip():
            self._log('⏭️ 跳过超时标记')
            dlg.destroy()
            self._do_generate(project_file, output_dir, template_file, {})

        tk.Button(btn_frame, text='✅ 确认并生成', font=('Microsoft YaHei', 12, 'bold'),
                  bg=C['accent'], fg='white', relief='flat', cursor='hand2',
                  padx=28, pady=8, activebackground=C['accent_a'],
                  command=_confirm).pack(side='left', padx=(0, 10))
        tk.Button(btn_frame, text='⏭️ 跳过（无超时）', font=('Microsoft YaHei', 10),
                  bg=C['orange'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=8, command=_skip).pack(side='left', padx=(0, 10))
        tk.Button(btn_frame, text='取消', font=('Microsoft YaHei', 10),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=8,
                  command=lambda: [dlg.destroy(), self._done()]).pack(side='right')

        dlg.protocol('WM_DELETE_WINDOW', lambda: [dlg.destroy(), self._done()])

    def _do_generate(self, project_file, output_dir, template_file, overtime_data):
        """实际执行生成：用 CLI 脚本生成提成表"""
        self.run_btn.configure(state='disabled', bg=C['gray'], text='⏳ 生成中...')
        self.st.configure(text='⏳ 正在计算绩效和生成表格...')
        self.progress.start(10)
        self._log('🔍 第二步：计算提成并生成表格...')
        self._log('—' * 50)

        self._output_files = {}
        self._current_overtime_map = {
            str(pid): set(episodes) for pid, episodes in overtime_data.items()
        }

        with tempfile.NamedTemporaryFile(
                mode='w', encoding='utf-8', suffix='.json',
                prefix='overtime_', dir=SCRIPT_DIR, delete=False) as temp_file:
            json.dump(overtime_data, temp_file, ensure_ascii=False, indent=2)
            overtime_file = temp_file.name

        gc, gc_sys = self._load_gc_module()
        _, template_date = gc.get_month_from_template(template_file)
        year_match = re.match(r'(\d{4})年', template_date)
        self._generation_year = int(year_match.group(1)) if year_match else None
        gc_sys.path.pop(0)

        def worker():
            try:
                env = os.environ.copy()
                env['PYTHONIOENCODING'] = 'utf-8'
                p = subprocess.Popen([PYTHON_EXE, CLI_SCRIPT,
                                      project_file, template_file,
                                      output_dir, '--overtime-file', overtime_file],
                                     cwd=SCRIPT_DIR,
                                     stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                                     stderr=subprocess.STDOUT, text=True,
                                     encoding='utf-8', errors='replace',
                                     env=env)
                try: p.stdin.write('\n'); p.stdin.flush()
                except: pass
                for ln in iter(p.stdout.readline, ''):
                    line = ln.strip()
                    if line.startswith('OUTPUT_EXCEL='):
                        self._output_files['excel'] = line.split('=', 1)[1]
                    elif line.startswith('OUTPUT_HTML='):
                        self._output_files['html'] = line.split('=', 1)[1]
                    self.root.after(0, lambda l=ln: self._filter_line(l))
                p.wait()
                self.root.after(0, self._done)
            except Exception as e:
                self.root.after(0, lambda: self._log(f'❌ 错误: {e}'))
                self.root.after(0, self._done)
            finally:
                try:
                    os.unlink(overtime_file)
                except OSError:
                    pass

        threading.Thread(target=worker, daemon=True).start()

    def _done(self):
        self.progress.stop()
        self.run_btn.configure(state='normal', bg=C['green'], text='▶  一键生成提成表')
        self._log('—' * 50)

        excel_path = self._output_files.get('excel', '')
        html_path = self._output_files.get('html', '')

        excel_ok = excel_path and os.path.exists(excel_path)
        html_ok = html_path and os.path.exists(html_path)

        if excel_ok or html_ok:
            self._log('🎉 全部完成！')
            self.st.configure(text='● 就绪 · 生成完成')

            # 自动备份
            if self.auto_backup.get() and HAS_FEATURES:
                try:
                    backed = backup_output(excel_path, html_path, BACKUP_DIR)
                    if backed:
                        self._log(f'💾 已备份 {len(backed)} 个文件到 backup/')
                except Exception as e:
                    self._log(f'⚠️ 备份失败: {e}')

            # 自动生成个人绩效卡片
            if excel_ok and HAS_FEATURES:
                try:
                    import pandas as pd
                    gc, _sys = self._load_gc_module()

                    df = pd.read_excel(self.project_file, header=None)
                    records, group_pids = gc.parse_projects(
                        df,
                        default_year=self._generation_year,
                        overtime_map=self._current_overtime_map,
                    )
                    cd = gc.compute_commission(records, group_pids)
                    card_paths = generate_person_cards(records, cd, CARDS_DIR)
                    if card_paths:
                        self._log(f'🃏 已生成 {len(card_paths)-1} 张个人绩效卡片 -> 个人绩效卡片/')
                        try: os.startfile(card_paths[0])
                        except: pass
                    _sys.path.pop(0)
                except Exception as e:
                    self._log(f'⚠️ 卡片生成跳过: {e}')

            # 打开文件
            if excel_ok:
                try: os.startfile(excel_path)
                except: pass
            if html_ok:
                try: os.startfile(html_path)
                except: pass
            messagebox.showinfo('完成', f'✅ 提成表、统计简报和仪表盘已生成完毕！\n\n📊 {os.path.basename(excel_path)}\n📈 {os.path.basename(html_path)}')
        else:
            self._log('❌ 生成失败：未找到输出文件')
            self.st.configure(text='❌ 生成失败 · 请查看日志')
            messagebox.showerror('失败', '❌ 生成失败！\n\n请检查日志了解详细错误信息。\n常见原因：\n  1. 项目数据文件格式不正确\n  2. 模板文件表头不匹配\n  3. config.json 人员配置有误')


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == '__main__':
    main()
